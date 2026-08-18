#!/usr/bin/env python3
"""
TEST VERSION: Single-run monitoring pipeline with NEW PROMPTS.
Outputs to TEST_OUTPUT directory for comparison with production.

This is identical to news_pipeline_single_run.py but with:
1. New category-specific expert prompts
2. Output to TEST_OUTPUT directory
3. Skip evaluation for 'other' category (cost savings)
"""

from __future__ import annotations

import argparse
import copy
from difflib import SequenceMatcher
import hashlib
import json
import logging
import os
import re
import shutil
import sys
import time
import warnings
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple
from zoneinfo import ZoneInfo

import jdatetime
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ==================== CONFIGURATION ====================
BASE_LIST_URL = "https://www.khabarfoori.com/بخش-اخبار-2"
DOMAIN = "https://www.khabarfoori.com"
HARD_MAX_PAGES = 500
STOP_AFTER_STALE_PAGES = 2
UNCERTAIN_DATE_EARLY_PAGE_LIMIT = 2
EXPECTED_LINKS_PER_PAGE = 10
LIST_PAGE_RETRY_COUNT = 2
REQUEST_TIMEOUT = 20
REQUEST_DELAY_SECONDS = 0.4
LOOP_INTERVAL_MINUTES = 30
TEHRAN_TZ = ZoneInfo("Asia/Tehran")

API_BASE_URL = "https://api.gapgpt.app/v1"
API_MODEL = "gemini-2.0-flash-lite"
API_KEY = os.getenv(
    "GAPGPT_API_KEY",
    "sk-c2hKxtnB7Ikdz0vtUrYKfeXSMIB2c6aU2DkELTMkDMyY0bXZ",
)

# The user asked to keep the prior, more specific rates.
INPUT_COST_PER_1M_TOKENS = 0.10
OUTPUT_COST_PER_1M_TOKENS = 0.40

SCHEMA_VERSION = 3
EVALUATION_VERSION = "2026-04-security-gold-v3"
COMPRESSION_VERSION = "2026-04-one-line-v3"
CLASSIFICATION_VERSION = "2026-04-llm-classification-v3"
CLASSIFICATION_PROMPT_MEMORY_VERSION = "2026-04-classification-memory-v3"
PROMPT_MEMORY_FILE = "prompt_memory.json"
COMPRESSION_TARGET_MAX_WORDS = 32
CLASSIFICATION_MEMORY_PROMPT_ENABLED = True
CLASSIFICATION_MEMORY_PROMPT_LIMIT = 64
CLASSIFICATION_MEMORY_PROMPT_CHAR_LIMIT = 12000
PROMPT_MEMORY_EXAMPLES_LIMIT = 40
PROMPT_MEMORY_EXAMPLES_CHAR_LIMIT = 5000
LEGACY_SUMMARY_TRUNCATED_CHAR_LENGTH = 200
LEGACY_SUMMARY_REBUILD_MAX_WORDS = 56

BASE_DIR = Path(__file__).resolve().parent
RUNTIME_DIR = BASE_DIR / "TEST_OUTPUT"
JSON_DIR = RUNTIME_DIR / "JSON Files"
LOG_DIR = RUNTIME_DIR / "LOG Files"
EXCEL_DIR = RUNTIME_DIR / "Excel Files"
MARKDOWN_DIR = RUNTIME_DIR / "Markdown Files"
TXT_DIR = RUNTIME_DIR / "TXT Files"
CLASSIFICATION_RUBRIC_FILE = MARKDOWN_DIR / "classification_rubric.md"
CLASSIFICATION_MEMORY_FILE = MARKDOWN_DIR / "classification_memory.md"
NEWS_DB_FILE = JSON_DIR / "news_database.json"
TRACKING_FILE = JSON_DIR / "news_tracking.json"
IMPORTANT_NEWS_FILE = RUNTIME_DIR / "important_news.txt"
API_COST_FILE = JSON_DIR / "api_costs.json"
OTHER_NEWS_FILE = JSON_DIR / "other_news_audit.json"
REFERENCE_WORKBOOK_FILE = Path("/Users/parham/Downloads/ثبت و تحلیل خبر.xlsx")
EXCEL_BASE_NAME = "ثبت و تحلیل خبر"
EXCEL_OUTPUT_FILE = EXCEL_DIR / f"{EXCEL_BASE_NAME}.xlsx"
PROMPT_MEMORY_PATH = JSON_DIR / PROMPT_MEMORY_FILE

HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36",
    "Accept-Language": "fa-IR,fa;q=0.9,en;q=0.8",
}

LEVELS = ["خیلی کم", "کم", "متوسط", "زیاد", "خیلی زیاد"]
TREND_LEVELS = ["↑", "↓", "→", "?", "خنثی", "نامطمئن"]

LEVEL_TO_SCORE = {
    "خیلی کم": 1,
    "کم": 2,
    "متوسط": 3,
    "زیاد": 4,
    "خیلی زیاد": 5,
}

EXCEL_HEADERS = [
    "شناسه خبر",
    "تاریخ انتشار",
    "ساعت انتشار",
    "منبع",
    "تیتر خبر",
    "اطمینان از وقوع خبر",
    "چقدر بر تغییر قیمت طلا اثر دارد؟",
    "چقدربا امنیت مرتبط است ؟",
    "جهت طلا",
    "وضعیت اطلاع رسانی",
    "توضیحات",
    "لینک",
]

WORKBOOK_NOTIFICATION_TRUE = "اطلاع‌رسانی شود"
WORKBOOK_NOTIFICATION_FALSE = "اطلاع‌رسانی نشود"

PERSIAN_MONTHS = {
    "فروردین": 1,
    "اردیبهشت": 2,
    "خرداد": 3,
    "تیر": 4,
    "مرداد": 5,
    "شهریور": 6,
    "مهر": 7,
    "آبان": 8,
    "آذر": 9,
    "دی": 10,
    "بهمن": 11,
    "اسفند": 12,
}

CLASSIFICATION_CATEGORIES = {"security", "economics", "security/economics", "other"}

ECONOMICS_KEYWORDS = [
    "دلار",
    "طلا",
    "نفت",
    "نرخ ارز",
    "نرخ بهره",
    "بانک مرکزی",
    "تورم",
    "سکه",
    "ریال",
    "حقوق پایه",
    "مالیات",
    "بودجه",
    "صادرات",
    "واردات",
    "یارانه",
    "قیمت‌گذاری",
    "قیمت گذاری",
    "ذخایر ارزی",
    "بازار آزاد",
    "نرخ رسمی",
    "نقدینگی",
    "پیش‌فروش",
    "پیش فروش",
    "مجوز واردات",
    "افزایش قیمت کالاهای اساسی",
    "کاهش مالیات",
    "کالابرگ",
    "تحریم",
    "عوارض",
    "تعرفه",
    "بورس کالا",
    "کنترل بانکی",
    "تسهیلات",
    "رمز ارز",
    "رمزارز",
]

SECURITY_KEYWORDS = [
    "جنگ",
    "صلح",
    "حمله",
    "تهدید",
    "درگیری",
    "انفجار",
    "بمب",
    "موشک",
    "پهپاد",
    "رزمایش",
    "توقیف نفتکش",
    "ناو",
    "تنگه هرمز",
    "حمله سایبری",
    "ترور",
    "آشوب",
    "ناآرامی",
    "امنیت مرز",
    "آتش‌بس",
    "آتش بس",
    "میانجی‌گری",
    "میانجی گری",
    "تحریم تسلیحاتی",
]

CLASSIFICATION_RUBRIC_MARKDOWN = """# Classification Rubric

## Allowed labels
- `security`
- `economics`
- `security/economics`
- `other`

## Mission
Your ONLY job is to classify the article into one of the four categories above.
Focus on the PRIMARY subject and MAIN meaning of the article.

Classify based on whether the article is relevant to:
- security concerns tied to Iran
- economics or financial conditions tied to Iran
- world-level macro or geopolitical events that can reasonably matter for Iran or global markets

## Decision Tree (Follow This Order)

1. **What is the PRIMARY subject?** Read the full article and identify the main topic.
2. **Does it have DIRECT relevance to Iran's security or economy?** If yes, classify accordingly.
3. **Does it have INDIRECT but MATERIAL relevance to Iran or global macro/security conditions?** If yes, classify accordingly.
4. **If neither direct nor material indirect relevance exists** → classify as `other`

## Critical Rules

- **Judge by MAIN MEANING, not isolated keywords.** Keywords are evidence, not the decision.
- **Weak overlap is not enough.** If the article only mentions security/economic terms in passing, classify as `other`.
- **Country-specific news with no Iran relevance** → classify as `other`
- **Novel or sudden important news should be classified fairly**, even if it differs from historical patterns.

## Economics coverage
Use `economics` when the core story is about macroeconomics, financial markets, trade, sanctions, energy markets, public finance, banking, inflation, exchange rates, taxes, subsidies, wages, supply pressure, or policy decisions that can affect Iran's economy or major global markets.

Examples of economics evidence:
دلار، طلا، نفت، نرخ ارز، نرخ بهره، بانک مرکزی، تورم، سکه، ریال، حقوق پایه، مالیات، بودجه، صادرات، واردات، یارانه، قیمت‌گذاری، ذخایر ارزی، بازار آزاد، نرخ رسمی، نقدینگی، پیش‌فروش، مجوز واردات، افزایش قیمت کالاهای اساسی، کاهش مالیات، کالابرگ، تحریم، عوارض، تعرفه، بورس کالا، کنترل بانکی، تسهیلات، رمز ارز

## Security coverage
Use `security` when the core story is about war, military action, terrorism, border security, cyberattack, sanctions with strategic-security consequences, regional escalation, threats, conflict management, or geopolitical instability that directly affects Iran or broad regional/global security.

Examples of security evidence:
جنگ، صلح، حمله، تهدید، درگیری، انفجار، بمب، موشک، پهپاد، رزمایش، توقیف نفتکش، ناو، تنگه هرمز، حمله سایبری، ترور، آشوب، ناآرامی، امنیت مرز، آتش‌بس، میانجی‌گری، تحریم تسلیحاتی

## Mixed label rule
Use `security/economics` only when both themes are materially central to the same article, not when one side is only a passing side-effect.
**Be careful:** Misclassifying into single category when it should be mixed can cause evaluation problems.

## Exclusions
Default to `other` for:
- local crime or local accidents without broader security meaning
- sports, entertainment, celebrity, lifestyle, culture, education, health, science, technology, religion, festivals
- general politics without clear security or economic stakes
- foreign country domestic news that is not materially relevant to Iran or major world macro/security conditions
- human-interest, social commentary, and viral content

## Positive examples
- Military strike on Iranian territory or forces: `security`
- Central bank policy, exchange-rate shock, inflation surge, oil-market disruption with macro effect: `economics`
- Strait of Hormuz escalation that also materially affects oil prices and trade risk: `security/economics`

## Negative examples
- Local municipal issue in another country with no Iran or broad macro/security significance: `other`
- Entertainment or sports story that mentions "crisis" or "attack" figuratively: `other`
- Technology or science news that mentions markets or risk casually: `other`
- Political rhetoric or statements without concrete security or economic implications: `other`
- Celebrity news, viral content, human interest stories: `other`
"""


# ==================== UI HELPERS ====================
class CliStyle:
    RESET = "\033[0m"
    BOLD = "\033[1m"
    DIM = "\033[2m"
    BLUE = "\033[94m"
    CYAN = "\033[96m"
    GREEN = "\033[92m"
    YELLOW = "\033[93m"
    RED = "\033[91m"
    MAGENTA = "\033[95m"


def colorize(text: str, color: str) -> str:
    if not sys.stdout.isatty():
        return text
    return f"{color}{text}{CliStyle.RESET}"


def print_box(title: str, rows: List[Tuple[str, str]]):
    width = 86
    title_text = f" {title} "
    print("╔" + "═" * width + "╗")
    print("║" + title_text.ljust(width) + "║")
    print("╟" + "─" * width + "╢")
    for key, value in rows:
        line = f"  {key:<24} {value}"
        print("║" + line.ljust(width) + "║")
    print("╚" + "═" * width + "╝")


def print_stage(stage: str, message: str):
    prefix = colorize(f"[{stage}]", CliStyle.CYAN)
    print(f"{prefix} {message}")


# ==================== DATA HELPERS ====================
def now_tehran() -> datetime:
    return datetime.now(TEHRAN_TZ)


def normalize_text(value: str) -> str:
    value = value or ""
    value = value.replace("\u200c", " ").replace("\xa0", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def compact_sentence(value: str) -> str:
    value = normalize_text(value).replace("\n", " ")
    value = re.sub(r"\s*([،,:;.!؟])\s*", r"\1 ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def normalize_category(value: str) -> str:
    normalized = normalize_text(value).lower()
    mapping = {
        "security": "security",
        "economics": "economics",
        "security/economics": "security/economics",
        "security-economics": "security/economics",
        "security_economics": "security/economics",
        "other": "other",
    }
    return mapping.get(normalized, "other")


def normalize_classification_confidence(value: str) -> str:
    normalized = normalize_text(value)
    if normalized in LEVELS:
        return normalized
    return "متوسط"


def normalize_memory_keywords(value: Any, fallback: Optional[List[str]] = None) -> List[str]:
    fallback = fallback or []
    if isinstance(value, str):
        candidates = re.split(r"[,|،\n]+", value)
    elif isinstance(value, list):
        candidates = value
    else:
        candidates = []

    keywords: List[str] = []
    for item in candidates:
        normalized = normalize_text(str(item))
        if not normalized or normalized in keywords:
            continue
        keywords.append(normalized)
        if len(keywords) >= 3:
            break

    for item in fallback:
        normalized = normalize_text(str(item))
        if not normalized or normalized in keywords:
            continue
        keywords.append(normalized)
        if len(keywords) >= 3:
            break
    return keywords


def normalize_memory_logic(value: str, fallback: str = "") -> str:
    logic = compact_sentence(value or fallback)
    words = logic.split()
    if len(words) > 12:
        logic = " ".join(words[:12])
    return logic


def keyword_hits(text: str, keywords: Iterable[str]) -> List[str]:
    hits: List[str] = []
    for keyword in keywords:
        pattern = rf"(?<![\w\u0600-\u06FF]){re.escape(keyword)}(?![\w\u0600-\u06FF])"
        if re.search(pattern, text):
            hits.append(keyword)
    return hits


def build_classification_evidence(article: Dict[str, Any]) -> Dict[str, Any]:
    combined_text = normalize_text(
        " ".join(
            [
                article.get("original_title", ""),
                article.get("lead", ""),
                article.get("content", ""),
            ]
        )
    )
    economics_hits = keyword_hits(combined_text, ECONOMICS_KEYWORDS)
    security_hits = keyword_hits(combined_text, SECURITY_KEYWORDS)

    if economics_hits and security_hits:
        suggested = "security/economics"
    elif security_hits:
        suggested = "security"
    elif economics_hits:
        suggested = "economics"
    else:
        suggested = "other"

    return {
        "economics_hits": economics_hits,
        "security_hits": security_hits,
        "economics_hit_count": len(economics_hits),
        "security_hit_count": len(security_hits),
        "has_both": bool(economics_hits and security_hits),
        "suggested_category": suggested,
    }


def persian_digits_to_english(text: str) -> str:
    return text.translate(str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789"))


def safe_move_artifact(source: Path, destination: Path):
    if not source.exists() or source.resolve() == destination.resolve():
        return
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists():
        return
    shutil.move(str(source), str(destination))


def cleanup_empty_dir(path: Path):
    try:
        path.rmdir()
    except OSError:
        pass


def migrate_runtime_layout():
    legacy_log_dir = RUNTIME_DIR / "logs"
    legacy_markdown_dir = RUNTIME_DIR / "markdown_evaluations"

    json_migrations = {
        RUNTIME_DIR / "news_database.json": NEWS_DB_FILE,
        RUNTIME_DIR / "news_tracking.json": TRACKING_FILE,
        RUNTIME_DIR / "api_costs.json": API_COST_FILE,
        RUNTIME_DIR / "prompt_memory.json": PROMPT_MEMORY_PATH,
        RUNTIME_DIR / "other_news_audit.json": OTHER_NEWS_FILE,
    }
    for source, destination in json_migrations.items():
        safe_move_artifact(source, destination)

    markdown_migrations = {
        RUNTIME_DIR / "classification_rubric.md": CLASSIFICATION_RUBRIC_FILE,
        RUNTIME_DIR / "classification_memory.md": CLASSIFICATION_MEMORY_FILE,
    }
    for source, destination in markdown_migrations.items():
        safe_move_artifact(source, destination)

    if legacy_log_dir.exists():
        for path in legacy_log_dir.iterdir():
            safe_move_artifact(path, LOG_DIR / path.name)
        cleanup_empty_dir(legacy_log_dir)

    if legacy_markdown_dir.exists():
        for path in legacy_markdown_dir.iterdir():
            safe_move_artifact(path, MARKDOWN_DIR / path.name)
        cleanup_empty_dir(legacy_markdown_dir)

    for path in list(RUNTIME_DIR.glob("*.md")):
        if path.name == IMPORTANT_NEWS_FILE.name:
            continue
        safe_move_artifact(path, MARKDOWN_DIR / path.name)

    for path in list(RUNTIME_DIR.glob("*.xlsx")):
        safe_move_artifact(path, EXCEL_DIR / path.name)

    for path in list(RUNTIME_DIR.glob("*.log")) + list(RUNTIME_DIR.glob("*.jsonl")):
        safe_move_artifact(path, LOG_DIR / path.name)


def ensure_runtime_layout():
    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    JSON_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    MARKDOWN_DIR.mkdir(parents=True, exist_ok=True)
    TXT_DIR.mkdir(parents=True, exist_ok=True)
    migrate_runtime_layout()
    CLASSIFICATION_RUBRIC_FILE.write_text(CLASSIFICATION_RUBRIC_MARKDOWN, encoding="utf-8")
    if not CLASSIFICATION_MEMORY_FILE.exists():
        CLASSIFICATION_MEMORY_FILE.write_text("# Classification Memory\n\n", encoding="utf-8")
    if not OTHER_NEWS_FILE.exists():
        with OTHER_NEWS_FILE.open("w", encoding="utf-8") as f:
            json.dump({"updated_at": now_tehran().isoformat(), "count": 0, "articles": []}, f, ensure_ascii=False, indent=2)


def article_hash(title: str, lead: str, content: str) -> str:
    raw = f"{normalize_text(title)}|{normalize_text(lead)}|{normalize_text(content)[:1500]}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def extract_khabarfoori_article_id(url: str) -> str:
    match = re.search(r"/(\d{5,})(?:[-/?]|$)", normalize_text(url))
    return match.group(1) if match else ""


def normalized_title_key(value: str) -> str:
    normalized = normalize_text(value).lower()
    normalized = re.sub(r"[^\w\u0600-\u06FF]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def article_identity_key(article: Dict[str, Any]) -> str:
    article_id = extract_khabarfoori_article_id(article.get("url", ""))
    if article_id:
        return f"id:{article_id}"

    published_date = article.get("published_at_persian") or article.get("listing_published_at_persian") or ""
    published_time = article.get("published_time") or article.get("listing_published_time") or ""
    title = (
        article.get("optimized_title")
        or article.get("original_title")
        or article.get("listing_title")
        or article.get("title")
        or ""
    )
    title_key = normalized_title_key(title)
    if published_date and published_time and title_key:
        title_hash = hashlib.md5(title_key.encode("utf-8")).hexdigest()[:12]
        return f"slot:{published_date}|{published_time}|{title_hash}"

    return f"url:{normalize_text(article.get('url', ''))}"


def preferred_article_sort_key(article: Dict[str, Any]) -> Tuple[int, int, int, int, str]:
    return (
        1 if article.get("notified") else 0,
        1 if article.get("processed") else 0,
        1 if article.get("compression_status") == "success" else 0,
        len(normalize_text(article.get("content", ""))),
        article.get("fetched_at", ""),
    )


def mark_article_duplicate(article: Dict[str, Any], primary_url: str):
    article["duplicate_of"] = primary_url
    article["processed"] = True
    article["classification_status"] = "duplicate"
    article["evaluation_status"] = "duplicate"
    article["compression_status"] = "duplicate"
    article["notification_status"] = WORKBOOK_NOTIFICATION_FALSE
    article["notified"] = False


def cleanup_database_duplicates(database: Dict[str, Any]) -> int:
    groups: Dict[str, List[Tuple[str, Dict[str, Any]]]] = {}
    for url, article in database.items():
        groups.setdefault(article_identity_key(article), []).append((url, article))

    duplicates_marked = 0
    for group in groups.values():
        if len(group) < 2:
            continue

        primary_url, primary_article = max(group, key=lambda item: preferred_article_sort_key(item[1]))
        primary_article["duplicate_of"] = None
        primary_article.setdefault("notification_status", WORKBOOK_NOTIFICATION_FALSE)
        database[primary_url] = primary_article

        for url, article in group:
            if url == primary_url:
                continue
            if article.get("duplicate_of") != primary_url:
                duplicates_marked += 1
            mark_article_duplicate(article, primary_url)
            database[url] = article

    return duplicates_marked


def unique_articles_for_output(articles: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    unique: Dict[str, Dict[str, Any]] = {}
    for article in articles:
        key = article_identity_key(article)
        current = unique.get(key)
        if not current or preferred_article_sort_key(article) > preferred_article_sort_key(current):
            unique[key] = article
    return list(unique.values())


def sort_database_items_for_save(database: Dict[str, Any]) -> Dict[str, Any]:
    ordered = sorted(database.items(), key=lambda item: article_recency_sort_tuple(item[1]), reverse=True)
    return {url: article for url, article in ordered}


def gregorian_to_persian_date_str(dt: datetime) -> str:
    pdate = jdatetime.date.fromgregorian(year=dt.year, month=dt.month, day=dt.day)
    return f"{pdate.year:04d}-{pdate.month:02d}-{pdate.day:02d}"


def persian_date_display(persian_date_str: str) -> str:
    try:
        year, month, day = map(int, persian_date_str.replace("/", "-").split("-"))
        month_name = jdatetime.date.j_months_fa[month - 1]
        return f"{day} {month_name}"
    except Exception:
        return persian_date_str


def new_classification_prompt(article, evidence, memory_text, rubric_text):
    """Improved classification prompt - more objective and cleaner."""
    return f"""
شما یک طبقه بند خبری هستید. وظیفه شما فقط تعیین دسته خبر است.
این مرحله طبقه بندی است، نه ارزیابی. شدت اثر، جهت طلا یا اهمیت خبر را تعیین نکنید.

راهنمای رسمی طبقه بندی:
{rubric_text}

قواعد حیاتی:
- فقط JSON معتبر برگردان.
- فقط یکی از این چهار دسته: security، economics، security/economics، other
- بر اساس معنای اصلی و موضوع اولیه خبر تصمیم بگیر، نه کلیدواژه های پراکنده.
- کلیدواژه ها شواهد هستند، نه تصمیم نهایی.
- اگر خبر جدید یا متفاوت از الگوهای قبلی است، مستقل و منصفانه طبقه بندی کن.
- حافظه فقط برای کالیبراسیون است. اگر با خبر فعلی ناسازگار است، نادیده بگیر.

حافظه پویای طبقه بندی (30 روز اخیر):
{memory_text}

شواهد کلیدواژه ای:
- پیشنهاد اولیه: {evidence['suggested_category']}
- تعداد کلیدواژه اقتصادی: {evidence['economics_hit_count']}
- کلیدواژه های اقتصادی: {", ".join(evidence['economics_hits']) or "ندارد"}
- تعداد کلیدواژه امنیتی: {evidence['security_hit_count']}
- کلیدواژه های امنیتی: {", ".join(evidence['security_hits']) or "ندارد"}

ورودی:
تیتر اصلی: {article.get('original_title', '')}
لید: {article.get('lead', '')}
متن خبر:
{article.get('content', '')[:2500]}

قالب خروجی دقیق:
{{
  "category": "security|economics|security/economics|other",
  "confidence": "خیلی کم|کم|متوسط|زیاد|خیلی زیاد",
  "matched_economics_keywords": ["..."],
  "matched_security_keywords": ["..."],
  "reason": "حداکثر 20 کلمه، خیلی کوتاه و فقط برای لاگ",
  "memory_keywords": ["..."],
  "memory_logic": "حداکثر 12 کلمه"
}}
""".strip()


def new_security_evaluation_prompt(article, memory_text):
    """Security-specific evaluation prompt with expert persona."""
    return f"""
شما یک تحلیلگر امنیتی برای ملت ایران هستید.
وظیفه شما: تعیین کنید این خبر چقدر بر ریسک های امنیتی ایران (مستقیم یا غیرمستقیم) اثر دارد.

نقش شما:
- تحلیلگر امنیت ملی ایران
- تمرکز بر تهدیدات نظامی، ژئوپلیتیک، و ثبات منطقه ای
- ارزیابی محافظه کارانه و واقع بینانه

{memory_text}

قواعد سخت:
- فقط JSON معتبر برگردان.
- فقط بر اساس متن همین خبر تصمیم بگیر.
- از بزرگنمایی، حدس، و امتیازدهی سخاوتمندانه خودداری کن.
- برای "زیاد" و "خیلی زیاد" باید متن به صورت روشن و مستقیم به رخداد مهم اشاره کند.

ورودی:
منبع: {article.get('source', '')}
دسته بندی: {article.get('classification_category', '')}
تیتر اصلی: {article.get('original_title', '')}
لید: {article.get('lead', '')}
متن خبر:
{article.get('content', '')[:2500]}

قالب خروجی دقیق:
{{
  "confidence_occurrence": "خیلی کم|کم|متوسط|زیاد|خیلی زیاد",
  "security_relevance": "خیلی کم|کم|متوسط|زیاد|خیلی زیاد",
  "gold_trend": "↑|↓|→|?",
  "rationale": "حداکثر 20 کلمه، خیلی کوتاه و فقط برای لاگ"
}}
""".strip()


def new_economics_evaluation_prompt(article, memory_text):
    """Economics-specific evaluation prompt with expert persona."""
    return f"""
شما یک تحلیلگر اقتصادی برای ملت ایران هستید.
وظیفه شما: تعیین کنید این خبر چقدر بر اقتصاد ایران و قیمت طلا (مستقیم یا غیرمستقیم) اثر دارد.

نقش شما:
- تحلیلگر اقتصاد کلان و بازارهای مالی ایران
- تمرکز بر قیمت طلا، نرخ ارز، تورم، و فشارهای اقتصادی
- ارزیابی محافظه کارانه و واقع بینانه

{memory_text}

قواعد سخت:
- فقط JSON معتبر برگردان.
- فقط بر اساس متن همین خبر تصمیم بگیر.
- از بزرگنمایی، حدس، و امتیازدهی سخاوتمندانه خودداری کن.
- برای "زیاد" و "خیلی زیاد" باید متن به صورت روشن و مستقیم به رخداد مهم اشاره کند.

ورودی:
منبع: {article.get('source', '')}
دسته بندی: {article.get('classification_category', '')}
تیتر اصلی: {article.get('original_title', '')}
لید: {article.get('lead', '')}
متن خبر:
{article.get('content', '')[:2500]}

قالب خروجی دقیق:
{{
  "confidence_occurrence": "خیلی کم|کم|متوسط|زیاد|خیلی زیاد",
  "gold_price_impact": "خیلی کم|کم|متوسط|زیاد|خیلی زیاد",
  "security_relevance": "خیلی کم|کم|متوسط|زیاد|خیلی زیاد",
  "gold_trend": "↑|↓|→|?",
  "rationale": "حداکثر 20 کلمه، خیلی کوتاه و فقط برای لاگ"
}}
""".strip()


def new_mixed_evaluation_prompt(article, memory_text):
    """Mixed security/economics evaluation prompt with strategic analyst persona."""
    return f"""
شما یک تحلیلگر استراتژیک برای ملت ایران هستید.
این خبر هم جنبه امنیتی دارد و هم اقتصادی.

نقش شما:
- تحلیلگر استراتژیک با دید ترکیبی امنیتی-اقتصادی
- تمرکز بر تعامل بین امنیت و اقتصاد
- ارزیابی محافظه کارانه و واقع بینانه

{memory_text}

قواعد سخت:
- فقط JSON معتبر برگردان.
- فقط بر اساس متن همین خبر تصمیم بگیر.
- از بزرگنمایی، حدس، و امتیازدهی سخاوتمندانه خودداری کن.
- برای "زیاد" و "خیلی زیاد" باید متن به صورت روشن و مستقیم به رخداد مهم اشاره کند.

ورودی:
منبع: {article.get('source', '')}
دسته بندی: {article.get('classification_category', '')}
تیتر اصلی: {article.get('original_title', '')}
لید: {article.get('lead', '')}
متن خبر:
{article.get('content', '')[:2500]}

قالب خروجی دقیق:
{{
  "confidence_occurrence": "خیلی کم|کم|متوسط|زیاد|خیلی زیاد",
  "gold_price_impact": "خیلی کم|کم|متوسط|زیاد|خیلی زیاد",
  "security_relevance": "خیلی کم|کم|متوسط|زیاد|خیلی زیاد",
  "gold_trend": "↑|↓|→|?",
  "rationale": "حداکثر 20 کلمه، خیلی کوتاه و فقط برای لاگ"
}}
""".strip()


def persian_date_display_full(persian_date_str: str) -> str:
    try:
        year, month, day = map(int, persian_date_str.replace("/", "-").split("-"))
        month_name = jdatetime.date.j_months_fa[month - 1]
        return f"{day} {month_name} {year}"
    except Exception:
        return persian_date_str


def trend_display_value(value: str) -> str:
    normalized = normalize_text(value)
    if normalized in ("→", "خنثی"):
        return "خنثی"
    if normalized in ("?", "نامطمئن"):
        return "نامطمئن"
    return normalized


def workbook_daily_path(persian_date_str: str) -> Path:
    return EXCEL_DIR / f"{EXCEL_BASE_NAME} - {persian_date_display_full(persian_date_str)}.xlsx"


def current_window_dates() -> Dict[str, str]:
    now = now_tehran()
    today_p = jdatetime.date.fromgregorian(year=now.year, month=now.month, day=now.day)
    yesterday_p = today_p - timedelta(days=1)
    return {
        "today": f"{today_p.year:04d}-{today_p.month:02d}-{today_p.day:02d}",
        "yesterday": f"{yesterday_p.year:04d}-{yesterday_p.month:02d}-{yesterday_p.day:02d}",
    }


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default


def save_json(path: Path, data: Any):
    ensure_runtime_layout()
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def estimate_tokens(text: str) -> int:
    return max(1, len(text) // 4)


def parse_iso_datetime(value: str) -> Optional[datetime]:
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            return dt.replace(tzinfo=TEHRAN_TZ)
        return dt.astimezone(TEHRAN_TZ)
    except Exception:
        return None


def parse_textual_persian_datetime(raw_text: str) -> Optional[datetime]:
    text = normalize_text(persian_digits_to_english(raw_text))
    if not text:
        return None

    slash_match = re.search(r"(?P<year>\d{4})[/-](?P<month>\d{1,2})[/-](?P<day>\d{1,2})", text)
    date_match = re.search(
        r"(?P<day>\d{1,2})\s+(?P<month>فروردین|اردیبهشت|خرداد|تیر|مرداد|شهریور|مهر|آبان|آذر|دی|بهمن|اسفند)\s+(?P<year>\d{4})",
        text,
    )
    time_match = re.search(r"(?P<hour>\d{1,2}):(?P<minute>\d{2})", text)

    if slash_match:
        year = int(slash_match.group("year"))
        month = int(slash_match.group("month"))
        day = int(slash_match.group("day"))
    elif date_match:
        day = int(date_match.group("day"))
        month = PERSIAN_MONTHS[date_match.group("month")]
        year = int(date_match.group("year"))
    else:
        return None
    hour = int(time_match.group("hour")) if time_match else 0
    minute = int(time_match.group("minute")) if time_match else 0

    try:
        greg = jdatetime.datetime(year, month, day, hour, minute).togregorian()
        return greg.replace(tzinfo=TEHRAN_TZ)
    except Exception:
        return None


def article_recency_datetime(article: Dict[str, Any]) -> Optional[datetime]:
    for field in (
        "published_at_gregorian",
        "listing_published_at_gregorian",
        "fetched_at",
    ):
        parsed = parse_iso_datetime(article.get(field, ""))
        if parsed:
            return parsed
    return None


def article_recency_sort_tuple(article: Dict[str, Any]) -> Tuple[str, str]:
    recency = article_recency_datetime(article)
    recency_key = recency.isoformat() if recency else ""
    return recency_key, article.get("url", "")


def extract_json_object(text: str) -> Dict[str, Any]:
    match = re.search(r"\{.*\}", text, re.DOTALL)
    payload = match.group(0) if match else text
    return json.loads(payload)


def similarity_ratio(left: str, right: str) -> float:
    return SequenceMatcher(None, normalize_text(left), normalize_text(right)).ratio()


def validate_level(value: str) -> str:
    if not value or value not in LEVELS:
        return "متوسط"
    return value


def validate_trend(value: str) -> str:
    return value if value in TREND_LEVELS else "?"


def calculate_notification_status(scores: Dict[str, str]) -> str:
    values = [
        LEVEL_TO_SCORE.get(scores.get("confidence", ""), 0),
        LEVEL_TO_SCORE.get(scores.get("gold_impact", ""), 0),
        LEVEL_TO_SCORE.get(scores.get("security_relevance", ""), 0),
    ]
    high_count = sum(1 for item in values if item >= 4)
    low_floor = min(values) >= 2
    return WORKBOOK_NOTIFICATION_TRUE if high_count >= 2 and low_floor else WORKBOOK_NOTIFICATION_FALSE


def build_human_news_entry(article: Dict[str, Any]) -> str:
    published = article.get("published_at_persian", "")
    published_display = persian_date_display_full(published) if published else ""
    time_text = article.get("published_time", "")
    source = article.get("source", "خبر فوری")
    meta = " | ".join(part for part in [source, published_display, time_text] if part)

    summary = article.get("one_line_description", "").strip()

    parts = [
        article.get("optimized_title", article.get("original_title", "")).strip(),
        "",
        summary,
        "",
        meta,
        "-" * 50,
        "",
    ]
    return "\n".join(parts)


def excel_notification_formula(row_num: int) -> str:
    metric_range = f"F{row_num}:H{row_num}"
    return (
        f'=IF(AND(COUNTIF({metric_range},"زیاد")+COUNTIF({metric_range},"خیلی زیاد")>=2,'
        f'COUNTIF({metric_range},"خیلی کم")=0),'
        f'"{WORKBOOK_NOTIFICATION_TRUE}","{WORKBOOK_NOTIFICATION_FALSE}")'
    )


def load_workbook_without_ooxml_extension_warnings(path: Path):
    """openpyxl emits UserWarning when stripping unsupported OOXML extensions from templates."""
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r".*extension is not supported and will be removed",
            category=UserWarning,
        )
        return load_workbook(path)


def template_sheet_assets() -> Dict[str, Any]:
    workbook = load_workbook_without_ooxml_extension_warnings(REFERENCE_WORKBOOK_FILE)
    sheet = workbook["بررسی خبر"]
    def snapshot(cell):
        return {
            "font": copy.copy(cell.font),
            "fill": copy.copy(cell.fill),
            "border": copy.copy(cell.border),
            "alignment": copy.copy(cell.alignment),
            "number_format": cell.number_format,
            "protection": copy.copy(cell.protection),
        }
    assets = {
        "column_widths": {
            "A": sheet.column_dimensions["A"].width,
            "B": 18.5,
            "C": sheet.column_dimensions["C"].width,
            "D": sheet.column_dimensions["D"].width,
            "E": sheet.column_dimensions["E"].width,
            "F": sheet.column_dimensions["F"].width,
            "G": sheet.column_dimensions["G"].width,
            "H": sheet.column_dimensions["H"].width,
            "I": sheet.column_dimensions["I"].width,
            "J": sheet.column_dimensions["J"].width,
            "K": sheet.column_dimensions["K"].width,
            "L": 37.5,
        },
        "row2_height": sheet.row_dimensions[2].height or 48,
        "header_styles": {coord: snapshot(sheet[f"{coord}2"]) for coord in list("ABCDEFGHIJK")},
        "body_styles": {coord: snapshot(sheet[f"{coord}3"]) for coord in list("ABCDEFGHIJK")},
    }
    workbook.close()
    return assets


def build_news_workbook(rows: List[Dict[str, Any]], target_path: Path):
    assets = template_sheet_assets()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "بررسی خبر"
    sheet.sheet_view.rightToLeft = True
    sheet.sheet_view.zoomScale = 120
    sheet.sheet_view.zoomScaleNormal = 120
    sheet.freeze_panes = "A3"
    sheet.sheet_properties.tabColor = "1F4E78"
    sheet.append([""] * len(EXCEL_HEADERS))
    sheet.append(EXCEL_HEADERS)
    sheet.row_dimensions[2].height = assets["row2_height"]

    for col_letter, width in assets["column_widths"].items():
        sheet.column_dimensions[col_letter].width = width

    for idx, col_letter in enumerate(list("ABCDEFGHIJK"), start=1):
        cell = sheet.cell(row=2, column=idx)
        style = assets["header_styles"][col_letter]
        cell.font = copy.copy(style["font"])
        cell.fill = copy.copy(style["fill"])
        cell.border = copy.copy(style["border"])
        cell.alignment = copy.copy(style["alignment"])
        cell.number_format = style["number_format"]
        cell.protection = copy.copy(style["protection"])
    link_header = sheet["L2"]
    style = assets["header_styles"]["K"]
    link_header.font = copy.copy(style["font"])
    link_header.fill = copy.copy(style["fill"])
    link_header.border = copy.copy(style["border"])
    link_header.alignment = copy.copy(style["alignment"])
    link_header.number_format = style["number_format"]
    link_header.protection = copy.copy(style["protection"])
    link_header.value = "لینک"

    notify_fill = PatternFill(fill_type="solid", fgColor="C3D69B")
    skip_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")

    for record_index, row_data in enumerate(rows, start=1):
        row_index = record_index + 2
        row_payload = dict(row_data)
        row_payload["شناسه خبر"] = record_index
        values = [row_payload.get(header, "") for header in EXCEL_HEADERS]
        sheet.append(values)
        for idx, col_letter in enumerate(list("ABCDEFGHIJK"), start=1):
            cell = sheet.cell(row=row_index, column=idx)
            style = assets["body_styles"][col_letter]
            cell.font = copy.copy(style["font"])
            cell.fill = copy.copy(style["fill"])
            cell.border = copy.copy(style["border"])
            cell.alignment = copy.copy(style["alignment"])
            cell.number_format = style["number_format"]
            cell.protection = copy.copy(style["protection"])
        link_cell = sheet.cell(row=row_index, column=12)
        style = assets["body_styles"]["E"]
        link_cell.fill = copy.copy(style["fill"])
        link_cell.border = copy.copy(style["border"])
        link_cell.number_format = style["number_format"]
        link_cell.protection = copy.copy(style["protection"])
        link_cell.alignment = Alignment(horizontal="right", vertical="center")
        link_cell.font = Font(name="Calibri", size=11, underline="single", color="0563C1")
        link_cell.value = row_data.get("لینک", "")
        if link_cell.value:
            link_cell.hyperlink = link_cell.value
        sheet.row_dimensions[row_index].height = 18

        computed_status = calculate_notification_status({
            "confidence": row_data.get("اطمینان از وقوع خبر", ""),
            "gold_impact": row_data.get("چقدر بر تغییر قیمت طلا اثر دارد؟", ""),
            "security_relevance": row_data.get("چقدربا امنیت مرتبط است ؟", ""),
        })
        is_notified = computed_status == WORKBOOK_NOTIFICATION_TRUE
        row_fill = notify_fill if is_notified else skip_fill
        for col_idx in range(1, 13):
            sheet.cell(row=row_index, column=col_idx).fill = copy.copy(row_fill)

        notification_cell = sheet.cell(row=row_index, column=10)
        notification_font = copy.copy(notification_cell.font)
        notification_font.bold = is_notified
        notification_cell.font = notification_font

    if sheet.max_row >= 3:
        rule_true = FormulaRule(
            formula=[f'$J3="{WORKBOOK_NOTIFICATION_TRUE}"'],
            stopIfTrue=False,
            fill=PatternFill(fill_type="solid", fgColor="C3D69B"),
        )
        rule_false = FormulaRule(
            formula=[f'$J3="{WORKBOOK_NOTIFICATION_FALSE}"'],
            stopIfTrue=False,
            fill=PatternFill(fill_type="solid", fgColor="D9D9D9"),
        )
        sheet.conditional_formatting.add(f"A3:L{sheet.max_row}", rule_true)
        sheet.conditional_formatting.add(f"A3:L{sheet.max_row}", rule_false)
        for row_idx in range(3, sheet.max_row + 1):
            sheet[f"J{row_idx}"] = excel_notification_formula(row_idx)

    sheet.auto_filter.ref = f"A2:L{max(2, sheet.max_row)}"

    workbook.save(target_path)


# ==================== LOGGING ====================
class CycleLogger:
    def __init__(self, cycle_id: str):
        ensure_runtime_layout()
        self.cycle_id = cycle_id
        self.text_path = LOG_DIR / f"{cycle_id}.log"
        self.jsonl_path = LOG_DIR / f"{cycle_id}.jsonl"
        self.logger = logging.getLogger(f"news_pipeline.{cycle_id}")
        self.logger.setLevel(logging.INFO)
        self.logger.handlers.clear()

        handler = logging.FileHandler(self.text_path, encoding="utf-8")
        handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
        self.logger.addHandler(handler)
        self.logger.propagate = False

    def info(self, message: str):
        self.logger.info(message)

    def warning(self, message: str):
        self.logger.warning(message)

    def error(self, message: str):
        self.logger.error(message)

    def event(self, category: str, event: str, **fields: Any):
        payload = {
            "ts": now_tehran().isoformat(),
            "cycle_id": self.cycle_id,
            "category": category,
            "event": event,
            **fields,
        }
        with self.jsonl_path.open("a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")


# ==================== SCRAPING ====================
def parse_listing_entries(html_content: str, logger: Optional[CycleLogger] = None) -> List[Dict[str, Any]]:
    soup = BeautifulSoup(html_content, "html.parser")
    container = soup.find("ul", class_="box container")
    if not container:
        return []

    entries: List[Dict[str, Any]] = []
    seen_keys: set[str] = set()

    for title_tag in container.find_all("h2", class_="title"):
        detail = title_tag.find_parent("div", class_="detail")
        time_tag = detail.find("time") if detail else None
        link_tag = title_tag.find("a", href=True)
        if not link_tag:
            continue

        href = normalize_text(link_tag.get("href", ""))
        if not href:
            continue

        if href.startswith("/"):
            full_url = f"{DOMAIN}{href}"
        elif href.startswith("http"):
            full_url = href
        else:
            full_url = f"{DOMAIN}/{href}"

        entry_key = article_identity_key({"url": full_url, "listing_title": link_tag.get_text(" ", strip=True)})
        if entry_key in seen_keys:
            continue
        seen_keys.add(entry_key)

        visible_title = normalize_text(link_tag.get_text(" ", strip=True))
        datetime_attr = normalize_text(time_tag.get("datetime", "")) if time_tag else ""
        time_span = time_tag.find("span") if time_tag else None
        visible_date_text = normalize_text(time_span.get_text(" ", strip=True) if time_span else "")

        visible_published_dt = parse_textual_persian_datetime(visible_date_text.replace("/", "-"))
        iso_published_dt = parse_iso_datetime(datetime_attr)

        if visible_published_dt:
            published_dt = visible_published_dt
        else:
            published_dt = iso_published_dt

        date_uncertain = published_dt is None
        if published_dt is None:
            published_dt = now_tehran()

        if logger and visible_published_dt and iso_published_dt:
            if gregorian_to_persian_date_str(visible_published_dt) != gregorian_to_persian_date_str(iso_published_dt):
                logger.event(
                    "date_filter",
                    "listing_timestamp_mismatch",
                    url=full_url,
                    listing_datetime=datetime_attr,
                    listing_visible_datetime=visible_date_text,
                    schema_persian=gregorian_to_persian_date_str(iso_published_dt),
                    visible_persian=gregorian_to_persian_date_str(visible_published_dt),
                )

        entries.append(
            {
                "url": full_url,
                "article_identity_key": entry_key,
                "listing_title": visible_title,
                "listing_datetime_attr": datetime_attr,
                "listing_date_text": visible_date_text,
                "listing_published_at_gregorian": published_dt.isoformat(),
                "listing_published_at_persian": gregorian_to_persian_date_str(published_dt),
                "listing_published_time": published_dt.strftime("%H:%M"),
                "date_uncertain": date_uncertain,
            }
        )

    return entries


def extract_article_data(
    session: requests.Session,
    url: str,
    logger: CycleLogger,
    listing_entry: Optional[Dict[str, Any]] = None,
) -> Optional[Dict[str, Any]]:
    try:
        response = session.get(url, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")

        title = normalize_text((listing_entry or {}).get("listing_title", "")) or "بدون تیتر"
        lead = ""
        content = ""
        source = "خبر فوری"
        datetime_attr = normalize_text((listing_entry or {}).get("listing_datetime_attr", ""))
        date_text = normalize_text((listing_entry or {}).get("listing_date_text", ""))

        for script in soup.find_all("script", type="application/ld+json"):
            try:
                raw = script.string
                if not raw:
                    continue
                data = json.loads(raw)
                if isinstance(data, dict) and data.get("@type") == "NewsArticle":
                    title = normalize_text(data.get("headline", title))
                    lead = normalize_text(data.get("description", lead))
                    datetime_attr = normalize_text(data.get("datePublished", datetime_attr))
                    break
            except json.JSONDecodeError:
                continue

        if title == "بدون تیتر":
            title_tag = soup.find("h1", class_="title")
            if title_tag:
                title = normalize_text(title_tag.get_text(" ", strip=True))

        lead_tag = soup.find("p", class_="lead")
        if lead_tag:
            lead = normalize_text(lead_tag.get_text(" ", strip=True))

        editor = soup.find("div", id="main_ck_editor")
        if editor:
            parts = []
            for p_tag in editor.find_all("p"):
                p_text = normalize_text(p_tag.get_text(" ", strip=True))
                if p_text:
                    parts.append(p_text)
            content = "\n".join(parts)

        time_container = soup.find("span", class_="news_time")
        if time_container:
            time_tag = time_container.find("time")
            if time_tag:
                date_text = normalize_text(time_tag.get_text(" ", strip=True))
                if not datetime_attr:
                    datetime_attr = normalize_text(time_tag.get("datetime", ""))

        source_div = soup.find("div", class_="source_news container")
        if source_div:
            source_title = source_div.find("span", class_="source_title")
            next_span = source_title.find_next_sibling("span") if source_title else None
            if next_span:
                source = normalize_text(next_span.get_text(" ", strip=True))

        iso_published_dt = parse_iso_datetime(datetime_attr)
        textual_published_dt = parse_textual_persian_datetime(date_text)

        # The site exposes two time sources. For the fetch window we trust the
        # on-page Persian timestamp first because it reflects the editorial
        # local publish date, while the schema timestamp can drift into the
        # next Persian day after timezone conversion.
        if iso_published_dt and textual_published_dt:
            if gregorian_to_persian_date_str(iso_published_dt) != gregorian_to_persian_date_str(textual_published_dt):
                logger.event(
                    "date_filter",
                    "timestamp_mismatch",
                    url=url,
                    schema_datetime=datetime_attr,
                    visible_datetime=date_text,
                    schema_persian=gregorian_to_persian_date_str(iso_published_dt),
                    visible_persian=gregorian_to_persian_date_str(textual_published_dt),
                )

        published_dt = textual_published_dt or iso_published_dt
        date_uncertain = published_dt is None
        if published_dt is None:
            published_dt = now_tehran()

        published_at_gregorian = published_dt.isoformat()
        published_at_persian = gregorian_to_persian_date_str(published_dt)
        published_time = published_dt.strftime("%H:%M")

        return {
            "url": url,
            "original_title": title,
            "lead": lead,
            "content": content,
            "source": source,
            "content_hash": article_hash(title, lead, content),
            "published_at_gregorian": published_at_gregorian,
            "published_at_persian": published_at_persian,
            "published_time": published_time,
            "date_uncertain": date_uncertain,
            "fetched_at": now_tehran().isoformat(),
            "listing_published_at_gregorian": (listing_entry or {}).get("listing_published_at_gregorian", published_at_gregorian),
            "listing_published_at_persian": (listing_entry or {}).get("listing_published_at_persian", published_at_persian),
        }
    except Exception as exc:
        logger.error(f"Article extraction failed for {url}: {exc}")
        logger.event("errors", "article_extraction_failed", url=url, error=str(exc))
        return None


def classification_memory_entry(article: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if article.get("duplicate_of"):
        return None
    if article.get("classification_status") != "success":
        return None
    if article.get("classification_method") == "rules_only":
        return None

    category = normalize_category(article.get("classification_category", "other"))
    keyword_hits = article.get("classification_keyword_hits", {})
    fallback_keywords = keyword_hits.get(category, []) if category in ("security", "economics") else []
    if category == "security/economics":
        fallback_keywords = keyword_hits.get("security", []) + keyword_hits.get("economics", [])

    keywords = normalize_memory_keywords(article.get("classification_memory_keywords"), fallback_keywords)
    logic = normalize_memory_logic(
        article.get("classification_memory_logic", ""),
        article.get("classification_rationale", ""),
    )
    if not keywords and not logic:
        return None

    return {
        "url": article.get("url", ""),
        "published_at_gregorian": article.get("published_at_gregorian", ""),
        "published_at_persian": article.get("published_at_persian", ""),
        "category": category,
        "keywords": keywords,
        "logic": logic,
        "title": normalize_text(article.get("original_title", ""))[:120],
    }


def classification_memory_entries(database: Dict[str, Any]) -> List[Dict[str, Any]]:
    entries = []
    for article in unique_articles_for_output(database.values()):
        entry = classification_memory_entry(article)
        if entry:
            recency_key, url_key = article_recency_sort_tuple(article)
            entry["_recency_sort"] = recency_key
            entry["_url_sort"] = url_key
            entries.append(entry)
    entries.sort(
        key=lambda item: (item.get("_recency_sort", ""), item.get("_url_sort", "")),
        reverse=True,
    )
    return entries


def render_classification_memory_markdown(entries: List[Dict[str, Any]]) -> str:
    lines = [
        "# Classification Memory",
        "",
        "خلاصه های کوتاه از منطق های قبلی طبقه بندی. این فایل از دیتابیس بازسازی می شود.",
        "",
    ]
    if not entries:
        lines.append("- هنوز نمونه ای ثبت نشده است.")
        return "\n".join(lines) + "\n"

    for idx, entry in enumerate(entries, start=1):
        keyword_text = "، ".join(entry["keywords"]) or "بدون کلیدواژه"
        title_text = entry["title"] or "بدون تیتر"
        lines.append(
            f"{idx}. [{entry['category']}] کلیدها: {keyword_text} | منطق: {entry['logic'] or 'بدون منطق'} | تیتر: {title_text}"
        )
    return "\n".join(lines) + "\n"


def prompt_memory_examples_text() -> str:
    payload = load_json(PROMPT_MEMORY_PATH, {"examples": []})
    examples = payload.get("examples", [])
    if not isinstance(examples, list) or not examples:
        return ""

    lines = ["## Historical Examples (JSON Memory)"]
    for item in examples[:PROMPT_MEMORY_EXAMPLES_LIMIT]:
        if not isinstance(item, dict):
            continue
        title = normalize_text(item.get("optimized_title", "")) or normalize_text(item.get("title", ""))
        category = normalize_category(item.get("classification_category", "other"))
        impact = normalize_text(item.get("gold_price_impact", "")) or "-"
        security = normalize_text(item.get("security_relevance", "")) or "-"
        line = f"- {category} | طلا: {impact} | امنیت: {security} | تیتر: {title[:90]}"
        candidate = "\n".join(lines + [line])
        if len(candidate) > PROMPT_MEMORY_EXAMPLES_CHAR_LIMIT:
            break
        lines.append(line)
    return "\n".join(lines) if len(lines) > 1 else ""


def filter_memory_by_time_window(entries: List[Dict[str, Any]], days: int = 30) -> List[Dict[str, Any]]:
    """Filter memory entries to only include those from the last N days."""
    cutoff = now_tehran() - timedelta(days=days)
    filtered = []
    
    for entry in entries:
        # Try to get the article's date
        article_date = entry.get("published_at_gregorian")
        if not article_date:
            article_date = entry.get("fetched_at")
        
        if article_date:
            try:
                dt = parse_iso_datetime(article_date)
                if dt and dt >= cutoff:
                    filtered.append(entry)
            except:
                # If we can't parse the date, include it to be safe
                filtered.append(entry)
        else:
            # If no date, include it
            filtered.append(entry)
    
    return filtered


def get_balanced_category_examples(entries: List[Dict[str, Any]], max_per_category: int = 15) -> List[Dict[str, Any]]:
    """Get balanced examples from each category to avoid bias."""
    categorized = {
        "security": [],
        "economics": [],
        "security/economics": [],
        "other": []
    }
    
    # Group by category
    for entry in entries:
        category = entry.get("category", "other")
        if category in categorized:
            categorized[category].append(entry)
    
    # Take up to max_per_category from each
    balanced = []
    for category in ["security", "economics", "security/economics", "other"]:
        balanced.extend(categorized[category][:max_per_category])
    
    return balanced


def get_category_specific_memory(entries: List[Dict[str, Any]], target_category: str, max_examples: int = 15) -> List[Dict[str, Any]]:
    """Get memory examples specific to a target category for evaluation."""
    filtered = [e for e in entries if e.get("category") == target_category]
    return filtered[:max_examples]


def classification_memory_prompt_text(database: Dict[str, Any]) -> str:
    if not CLASSIFICATION_MEMORY_PROMPT_ENABLED:
        return "حافظه پویا برای این درخواست غیرفعال است."

    # Get time-windowed entries (last 30 days)
    all_entries = classification_memory_entries(database)
    time_windowed_entries = filter_memory_by_time_window(all_entries, days=30)
    
    if not time_windowed_entries:
        return "نمونه قبلی برای حافظه طبقه بندی موجود نیست."

    # Get balanced examples from each category (10-20 per category)
    balanced_entries = get_balanced_category_examples(time_windowed_entries, max_per_category=15)
    
    lines = [
        "## Dynamic Classification Memory (Last 30 Days)",
        "این حافظه فقط برای کالیبراسیون است، نه تقلید.",
        "اگر خبر فعلی جدید یا متفاوت است، مستقل و منصفانه تصمیم بگیر.",
    ]
    for entry in balanced_entries:
        line = (
            f"- [{entry['category']}] کلیدها: {', '.join(entry['keywords']) or 'ندارد'} | "
            f"منطق: {entry['logic'] or 'ندارد'}"
        )
        candidate = "\n".join(lines + [line])
        if len(candidate) > CLASSIFICATION_MEMORY_PROMPT_CHAR_LIMIT:
            break
        lines.append(line)

    return "\n".join(lines)


def evaluation_memory_prompt_text(database: Dict[str, Any]) -> str:
    """Get evaluation memory filtered by time window (generic, not category-specific)."""
    all_entries = classification_memory_entries(database)
    time_windowed_entries = filter_memory_by_time_window(all_entries, days=30)
    
    # Filter out 'other' category for evaluation memory
    filtered = [e for e in time_windowed_entries if e.get("category") != "other"]
    
    if not filtered:
        return "نمونه قبلی برای حافظه ارزیابی موجود نیست."
    
    lines = [
        "## Evaluation Memory (Last 30 Days)",
        "این حافظه فقط برای کالیبراسیون است. اگر خبر فعلی جدید یا متفاوت است، مستقل تصمیم بگیر.",
    ]
    
    for entry in filtered[:20]:
        line = (
            f"- [{entry['category']}] منطق: {entry['logic'] or 'ندارد'} | "
            f"تیتر: {entry['title'][:60] if entry.get('title') else 'بدون تیتر'}"
        )
        candidate = "\n".join(lines + [line])
        if len(candidate) > 3200:
            break
        lines.append(line)
    
    return "\n".join(lines)


def category_specific_evaluation_memory(database: Dict[str, Any], category: str) -> str:
    """Get evaluation memory specific to a category."""
    all_entries = classification_memory_entries(database)
    time_windowed_entries = filter_memory_by_time_window(all_entries, days=30)
    
    # Get category-specific examples
    category_entries = get_category_specific_memory(time_windowed_entries, category, max_examples=15)
    
    if not category_entries:
        return f"نمونه قبلی برای دسته {category} موجود نیست."
    
    lines = [
        f"## {category.upper()} Evaluation Memory (Last 30 Days)",
        "این حافظه فقط برای کالیبراسیون است. خبر فعلی را مستقل ارزیابی کن.",
    ]
    
    for entry in category_entries:
        line = f"- منطق: {entry['logic'] or 'ندارد'} | تیتر: {entry['title'][:60] if entry.get('title') else 'بدون تیتر'}"
        candidate = "\n".join(lines + [line])
        if len(candidate) > 3200:
            break
        lines.append(line)
    
    return "\n".join(lines)


def rebuild_classification_memory_file(database: Dict[str, Any]) -> None:
    ensure_runtime_layout()
    CLASSIFICATION_MEMORY_FILE.write_text(
        render_classification_memory_markdown(classification_memory_entries(database)),
        encoding="utf-8",
    )


def rebuild_other_news_audit(database: Dict[str, Any]) -> None:
    ensure_runtime_layout()
    articles = []
    for article in unique_articles_for_output(database.values()):
        if article.get("duplicate_of"):
            continue
        if normalize_category(article.get("classification_category", "")) != "other":
            continue
        article_copy = copy.deepcopy(article)
        article_copy["classification_category"] = "other"
        articles.append(article_copy)

    articles.sort(key=article_recency_sort_tuple, reverse=True)
    with OTHER_NEWS_FILE.open("w", encoding="utf-8") as f:
        json.dump(
            {
                "updated_at": now_tehran().isoformat(),
                "count": len(articles),
                "articles": articles,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )


def backfill_classification_memory_fields(database: Dict[str, Any]) -> None:
    for article in database.values():
        if article.get("duplicate_of"):
            continue
        if article.get("classification_status") != "success":
            continue
        if article.get("classification_method") == "rules_only":
            continue
        entry = classification_memory_entry(article)
        if not entry:
            continue
        if article.get("classification_memory_keywords") in (None, []):
            article["classification_memory_keywords"] = entry["keywords"]
        if not article.get("classification_memory_logic"):
            article["classification_memory_logic"] = entry["logic"]
        if not article.get("classification_prompt_memory_version"):
            article["classification_prompt_memory_version"] = CLASSIFICATION_PROMPT_MEMORY_VERSION


def rebuild_prompt_memory_file(database: Dict[str, Any]) -> None:
    ensure_runtime_layout()
    entries = []
    for article in sorted(unique_articles_for_output(database.values()), key=article_recency_sort_tuple, reverse=True):
        if article.get("duplicate_of"):
            continue
        if article.get("classification_status") != "success":
            continue
        entries.append(
            {
                "published_at_gregorian": article.get("published_at_gregorian", ""),
                "classification_category": normalize_category(article.get("classification_category", "other")),
                "optimized_title": normalize_text(article.get("optimized_title", "")),
                "one_line_description": normalize_text(article.get("one_line_description", "")),
                "confidence_occurrence": normalize_text(article.get("confidence_occurrence", "")),
                "gold_price_impact": normalize_text(article.get("gold_price_impact", "")),
                "security_relevance": normalize_text(article.get("security_relevance", "")),
                "gold_trend": normalize_text(article.get("gold_trend", "")),
                "notification_status": normalize_text(article.get("notification_status", "")),
            }
        )
        if len(entries) >= PROMPT_MEMORY_EXAMPLES_LIMIT:
            break
    save_json(PROMPT_MEMORY_PATH, {"updated_at": now_tehran().isoformat(), "examples": entries})


def summary_needs_local_rebuild(summary: str) -> bool:
    text = compact_sentence(summary)
    if not text:
        return True
    if len(text) < LEGACY_SUMMARY_TRUNCATED_CHAR_LENGTH:
        return False
    if re.search(r"[.!؟]$", text):
        return False
    if text.endswith((" که", " و", " یا", " از", " با", " در", " به", " برای", " تا", " بر")):
        return True
    return True


def build_local_summary_from_article(article: Dict[str, Any]) -> str:
    parts = [normalize_text(article.get("lead", "")), normalize_text(article.get("content", ""))]
    base_text = normalize_text(" ".join(part for part in parts if part))
    if not base_text:
        return ""

    sentences = [normalize_text(s) for s in re.split(r"(?<=[.!؟])\s+", base_text) if normalize_text(s)]
    if not sentences:
        sentences = [base_text]

    summary = sentences[0]
    if len(summary.split()) < 14 and len(sentences) > 1:
        summary = f"{summary} {sentences[1]}"

    words = summary.split()
    if len(words) > LEGACY_SUMMARY_REBUILD_MAX_WORDS:
        summary = " ".join(words[:LEGACY_SUMMARY_REBUILD_MAX_WORDS]).rstrip("،,:;-")
    if not re.search(r"[.!؟]$", summary):
        summary += "."

    return compact_sentence(summary)


def repair_legacy_summaries(database: Dict[str, Any]) -> int:
    repaired = 0
    for article in database.values():
        if article.get("duplicate_of"):
            continue
        current = article.get("one_line_description", "")
        if not summary_needs_local_rebuild(current):
            continue
        rebuilt = build_local_summary_from_article(article)
        if rebuilt:
            article["one_line_description"] = rebuilt
            repaired += 1
    return repaired


# ==================== COST TRACKING ====================
COST_RUN_HISTORY_LIMIT = 400


def human_cost_summary(bucket: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "calls": int(bucket.get("calls", 0)),
        "input_tokens": int(bucket.get("input_tokens", 0)),
        "output_tokens": int(bucket.get("output_tokens", 0)),
        "total_tokens": int(bucket.get("total_tokens", 0)),
        "total_cost_usd": round(float(bucket.get("total_cost", 0.0)), 8),
        "display": {
            "total_cost_usd": f"${float(bucket.get('total_cost', 0.0)):.6f}",
            "input_tokens": f"{int(bucket.get('input_tokens', 0)):,}",
            "output_tokens": f"{int(bucket.get('output_tokens', 0)):,}",
            "total_tokens": f"{int(bucket.get('total_tokens', 0)):,}",
        },
    }


def serialize_stage_breakdown(stage_data: Dict[str, Any]) -> Dict[str, Any]:
    payload: Dict[str, Any] = {}
    for stage, bucket in sorted(stage_data.items()):
        payload[stage] = {
            "calls": int(bucket.get("calls", 0)),
            "input_tokens": int(bucket.get("input_tokens", 0)),
            "output_tokens": int(bucket.get("output_tokens", 0)),
            "total_tokens": int(bucket.get("total_tokens", 0)),
            "total_cost_usd": round(float(bucket.get("total_cost", 0.0)), 8),
        }
    return payload


def deserialize_stage_breakdown(stage_data: Dict[str, Any]) -> Dict[str, Any]:
    payload: Dict[str, Any] = {}
    for stage, bucket in (stage_data or {}).items():
        payload[stage] = {
            "calls": int(bucket.get("calls", 0)),
            "input_tokens": int(bucket.get("input_tokens", 0)),
            "output_tokens": int(bucket.get("output_tokens", 0)),
            "total_tokens": int(bucket.get("total_tokens", 0)),
            "total_cost": float(bucket.get("total_cost_usd", bucket.get("total_cost", 0.0))),
        }
    return payload


def default_cost_state() -> Dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "pricing": {
            "input_cost_per_1m_tokens": INPUT_COST_PER_1M_TOKENS,
            "output_cost_per_1m_tokens": OUTPUT_COST_PER_1M_TOKENS,
        },
        "lifetime": {
            "calls": 0,
            "input_tokens": 0,
            "output_tokens": 0,
            "total_tokens": 0,
            "total_cost": 0.0,
            "stages": {},
        },
        "cycle": {
            "calls": 0,
            "input_tokens": 0,
            "output_tokens": 0,
            "total_tokens": 0,
            "total_cost": 0.0,
            "stages": {},
        },
        "last_cycle": {},
        "runs": [],
    }


def merge_stage_cost(costs: Dict[str, Any], stage: str, input_tokens: int, output_tokens: int):
    input_cost = (input_tokens / 1_000_000) * INPUT_COST_PER_1M_TOKENS
    output_cost = (output_tokens / 1_000_000) * OUTPUT_COST_PER_1M_TOKENS
    total_cost = input_cost + output_cost

    for scope in ("lifetime", "cycle"):
        bucket = costs[scope]
        bucket["calls"] = bucket.get("calls", 0) + 1
        bucket["input_tokens"] = bucket.get("input_tokens", 0) + input_tokens
        bucket["output_tokens"] = bucket.get("output_tokens", 0) + output_tokens
        bucket["total_tokens"] = bucket.get("total_tokens", 0) + input_tokens + output_tokens
        bucket["total_cost"] = round(bucket.get("total_cost", 0.0) + total_cost, 8)
        stage_bucket = bucket.setdefault("stages", {}).setdefault(
            stage,
            {"calls": 0, "input_tokens": 0, "output_tokens": 0, "total_tokens": 0, "total_cost": 0.0},
        )
        stage_bucket["calls"] += 1
        stage_bucket["input_tokens"] += input_tokens
        stage_bucket["output_tokens"] += output_tokens
        stage_bucket["total_tokens"] += input_tokens + output_tokens
        stage_bucket["total_cost"] = round(stage_bucket["total_cost"] + total_cost, 8)


def _cost_run_from_bucket(bucket: Dict[str, Any], mode: str, force: Optional[bool] = None) -> Dict[str, Any]:
    run = {
        "id": bucket.get("cycle_id"),
        "mode": mode,
        "model": API_MODEL,
        "started_at": bucket.get("started_at"),
        "finished_at": bucket.get("finished_at"),
        "calls": int(bucket.get("calls", 0)),
        "tok_in": int(bucket.get("input_tokens", 0)),
        "tok_out": int(bucket.get("output_tokens", 0)),
        "tok_total": int(bucket.get("total_tokens", 0)),
        "usd_total": round(float(bucket.get("total_cost", 0.0)), 8),
        "stages": {},
    }
    if force is not None:
        run["force"] = bool(force)

    for stage, stage_bucket in sorted((bucket.get("stages") or {}).items()):
        run["stages"][stage] = {
            "calls": int(stage_bucket.get("calls", 0)),
            "tok_in": int(stage_bucket.get("input_tokens", 0)),
            "tok_out": int(stage_bucket.get("output_tokens", 0)),
            "tok_total": int(stage_bucket.get("total_tokens", 0)),
            "usd_total": round(float(stage_bucket.get("total_cost", 0.0)), 8),
        }
    return run


def append_cost_run(costs: Dict[str, Any], mode: str, force: Optional[bool] = None) -> None:
    bucket = costs.get("last_cycle") or costs.get("cycle") or {}
    run_id = bucket.get("cycle_id")
    if not run_id:
        return

    run = _cost_run_from_bucket(bucket, mode=mode, force=force)
    runs = costs.setdefault("runs", [])
    runs = [item for item in runs if item.get("id") != run_id]
    runs.append(run)
    costs["runs"] = sorted(runs, key=lambda item: item.get("started_at", ""), reverse=True)[:COST_RUN_HISTORY_LIMIT]


def load_cost_state_from_disk(raw: Dict[str, Any]) -> Dict[str, Any]:
    state = default_cost_state()
    if not raw:
        return state

    # Internal format (older and current runtime state)
    if "lifetime" in raw and "cycle" in raw:
        state.update(raw)
        state.setdefault("runs", [])
        return state

    # New concise persisted format
    if "meta" in raw and "runs" in raw:
        meta = raw.get("meta", {})
        totals = meta.get("totals", {})
        state["lifetime"].update(
            {
                "calls": int(totals.get("calls", 0)),
                "input_tokens": int(totals.get("tok_in", 0)),
                "output_tokens": int(totals.get("tok_out", 0)),
                "total_tokens": int(totals.get("tok_total", 0)),
                "total_cost": float(totals.get("usd_total", 0.0)),
            }
        )
        runs = raw.get("runs", []) if isinstance(raw.get("runs"), list) else []
        state["runs"] = runs[:COST_RUN_HISTORY_LIMIT]
        if state["runs"]:
            latest = state["runs"][0]
            state["last_cycle"] = {
                "cycle_id": latest.get("id"),
                "started_at": latest.get("started_at"),
                "finished_at": latest.get("finished_at"),
                "calls": int(latest.get("calls", 0)),
                "input_tokens": int(latest.get("tok_in", 0)),
                "output_tokens": int(latest.get("tok_out", 0)),
                "total_tokens": int(latest.get("tok_total", 0)),
                "total_cost": float(latest.get("usd_total", 0.0)),
                "stages": {
                    stage: {
                        "calls": int(info.get("calls", 0)),
                        "input_tokens": int(info.get("tok_in", 0)),
                        "output_tokens": int(info.get("tok_out", 0)),
                        "total_tokens": int(info.get("tok_total", 0)),
                        "total_cost": float(info.get("usd_total", 0.0)),
                    }
                    for stage, info in (latest.get("stages", {}) or {}).items()
                },
            }
        return state

    # Legacy persisted format
    lifetime = raw.get("lifetime_summary", {})
    last_cycle = raw.get("last_cycle_summary", {})
    stages = raw.get("stage_breakdown", {})

    for source, target in ((lifetime, "lifetime"), (last_cycle, "last_cycle")):
        if not source:
            continue
        state[target].update(
            {
                "calls": int(source.get("calls", 0)),
                "input_tokens": int(source.get("input_tokens", 0)),
                "output_tokens": int(source.get("output_tokens", 0)),
                "total_tokens": int(source.get("total_tokens", 0)),
                "total_cost": float(source.get("total_cost_usd", source.get("total_cost", 0.0))),
            }
        )
    state["lifetime"]["stages"] = deserialize_stage_breakdown(stages.get("lifetime", {}))
    state["last_cycle"]["stages"] = deserialize_stage_breakdown(stages.get("last_cycle", {}))
    return state


def cost_state_for_save(costs: Dict[str, Any]) -> Dict[str, Any]:
    runs = sorted(costs.get("runs", []), key=lambda item: item.get("started_at", ""), reverse=True)[:COST_RUN_HISTORY_LIMIT]
    latest = runs[0] if runs else {}
    lifetime = costs.get("lifetime", {})
    return {
        "meta": {
            "schema_version": SCHEMA_VERSION,
            "model": API_MODEL,
            "pricing": {
                "input_cost_per_1m_tokens": INPUT_COST_PER_1M_TOKENS,
                "output_cost_per_1m_tokens": OUTPUT_COST_PER_1M_TOKENS,
            },
            "totals": {
                "runs": len(runs),
                "calls": int(lifetime.get("calls", 0)),
                "tok_in": int(lifetime.get("input_tokens", 0)),
                "tok_out": int(lifetime.get("output_tokens", 0)),
                "tok_total": int(lifetime.get("total_tokens", 0)),
                "usd_total": round(float(lifetime.get("total_cost", 0.0)), 8),
            },
            "latest": {
                "id": latest.get("id"),
                "mode": latest.get("mode"),
                "finished_at": latest.get("finished_at"),
                "calls": int(latest.get("calls", 0) or 0),
                "usd_total": round(float(latest.get("usd_total", 0.0) or 0.0), 8),
            },
            "updated_at": now_tehran().isoformat(),
        },
        "runs": runs,
    }


# ==================== PIPELINE ====================
@dataclass
class CycleStats:
    cycle_id: str
    pages_scanned: int = 0
    links_seen: int = 0
    new_articles: int = 0
    duplicates: int = 0
    skipped_out_of_window: int = 0
    uncertain_included: int = 0
    classified: int = 0
    evaluated: int = 0
    compressed: int = 0
    notifying: int = 0
    skipped_other: int = 0
    security_count: int = 0
    economics_count: int = 0
    security_economics_count: int = 0
    other_count: int = 0
    failed: int = 0


class NewsPipeline:
    def __init__(self):
        ensure_runtime_layout()
        self.session = requests.Session()
        self.session.headers.update(HTTP_HEADERS)
        self.window = current_window_dates()

    def load_state(self) -> Tuple[Dict[str, Any], Dict[str, Any], Dict[str, Any]]:
        database = load_json(NEWS_DB_FILE, {})
        cleanup_database_duplicates(database)
        backfill_classification_memory_fields(database)
        repair_legacy_summaries(database)
        tracking = load_json(
            TRACKING_FILE,
            {
                "schema_version": SCHEMA_VERSION,
                "last_successful_cycle": None,
                "last_started_cycle": None,
                "cycles": [],
            },
        )
        costs = load_cost_state_from_disk(load_json(API_COST_FILE, default_cost_state()))
        rebuild_classification_memory_file(database)
        rebuild_prompt_memory_file(database)
        rebuild_other_news_audit(database)
        return database, tracking, costs

    def save_state(
        self,
        database: Dict[str, Any],
        tracking: Dict[str, Any],
        costs: Dict[str, Any],
        rebuild_auxiliary: bool = True,
    ):
        if rebuild_auxiliary:
            rebuild_classification_memory_file(database)
            rebuild_prompt_memory_file(database)
            rebuild_other_news_audit(database)
            self.rebuild_txt_files(database)
        save_json(NEWS_DB_FILE, sort_database_items_for_save(database))
        save_json(TRACKING_FILE, tracking)
        save_json(API_COST_FILE, cost_state_for_save(costs))

    def call_llm_json(self, prompt: str, stage: str, costs: Dict[str, Any], logger: CycleLogger) -> Dict[str, Any]:
        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json",
        }
        payload = {
            "model": API_MODEL,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.05,
            "max_tokens": 900,
        }

        retries = 2
        for attempt in range(1, retries + 2):
            response = self.session.post(
                f"{API_BASE_URL}/chat/completions",
                headers=headers,
                json=payload,
                timeout=REQUEST_TIMEOUT + 20,
            )
            response.raise_for_status()
            body = response.json()
            message_content = body["choices"][0]["message"]["content"]
            input_tokens = estimate_tokens(prompt)
            output_tokens = estimate_tokens(message_content)
            merge_stage_cost(costs, stage, input_tokens, output_tokens)
            try:
                return extract_json_object(message_content)
            except Exception as exc:
                logger.warning(f"Malformed {stage} JSON on attempt {attempt}: {exc}")
                logger.event("llm", "malformed_json", stage=stage, attempt=attempt, response=message_content[:800])
                payload["temperature"] = 0.0
                if attempt == retries + 1:
                    raise
        raise RuntimeError(f"{stage} failed after retries")

    def call_llm_text(self, prompt: str, stage: str, costs: Dict[str, Any], logger: CycleLogger) -> str:
        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json",
        }
        payload = {
            "model": API_MODEL,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.0,
            "max_tokens": 768 if stage == "compression" else 500,
        }
        response = self.session.post(
            f"{API_BASE_URL}/chat/completions",
            headers=headers,
            json=payload,
            timeout=REQUEST_TIMEOUT + 20,
        )
        response.raise_for_status()
        body = response.json()
        text = body["choices"][0]["message"]["content"]
        if stage == "compression" and body["choices"][0].get("finish_reason") == "length":
            logger.event("llm", "compression_truncated_by_api", url="", preview=text[:200])
        merge_stage_cost(costs, stage, estimate_tokens(prompt), estimate_tokens(text))
        logger.event("llm", "text_response", stage=stage, preview=text[:250])
        return text

    def classification_prompt(self, article: Dict[str, Any], evidence: Dict[str, Any], memory_text: str) -> str:
        rubric_text = CLASSIFICATION_RUBRIC_FILE.read_text(encoding="utf-8")
        # Use new classification prompt
        return new_classification_prompt(article, evidence, memory_text, rubric_text)

    def evaluation_prompt(self, article: Dict[str, Any], memory_text: str) -> str:
        # Use category-specific evaluation prompts
        category = article.get('classification_category', 'other')
        
        if category == "security":
            return new_security_evaluation_prompt(article, memory_text)
        elif category == "economics":
            return new_economics_evaluation_prompt(article, memory_text)
        elif category == "security/economics":
            return new_mixed_evaluation_prompt(article, memory_text)
        else:
            # For 'other' category, return None to skip evaluation
            return None

    def compression_prompt(self, article: Dict[str, Any]) -> str:
        return f"""
نقش شما: سیستم فشرده سازی و خلاصه سازی اخبار.

ورودی شامل «تیتر خبر» و «متن کامل خبر» است. وظیفه شما تولید یک «تیتر بهینه شده» و یک «شرح بسیار فشرده از کل خبر» است.

قوانین عمومی:
1. خروجی باید کاملا بی طرف و خبری باشد.
2. هیچ تحلیل، نظر شخصی، نتیجه گیری یا تفسیر اضافه نکن.
3. فقط اطلاعات موجود در متن خبر را فشرده و بازنویسی کن.
4. از اضافه کردن اطلاعات خارج از متن خودداری کن.
5. از کلمات اضافی، توضیحات غیرضروری و جملات تزئینی اجتناب کن.
6. لحن و چارچوب خبر را حفظ کن اما آن را شدیدتر، نرم تر یا ایدئولوژیک تر نکن.

الگوریتم خلاصه سازی:
ابتدا در متن خبر این موارد را تشخیص بده:
- رویداد اصلی
- افراد یا نهادهای اصلی
- مکان یا زمینه رویداد در صورت وجود
- مهم ترین نتیجه، تصمیم، یا پیامد

سپس خروجی فشرده را فقط بر اساس همین عناصر تولید کن.

قوانین مربوط به تیتر:
1. تیتر باید موضوع خبر را واضح و خبری بیان کند.
2. تیتر باید از نظر مفهومی بسیار نزدیک به تیتر اصلی باشد.
3. فقط در صورت نیاز برای وضوح، فشرده سازی، یا رفع ابهام آن را بازنویسی کن.
4. از تغییر شدید در مفهوم تیتر اصلی خودداری کن.
5. حق نداری اطلاعات جدید، قضاوت، یا نتیجه گیری اضافه کنی.
6. طول تیتر: 7 تا 12 کلمه.

قوانین مربوط به شرح:
1. شرح باید چکیده کامل خبر باشد.
2. خواننده شرح باید تقریباً اطلاعات اصلی خواننده متن کامل را دریافت کند.
3. شرح باید شامل مهم ترین عناصر خبر باشد: چه اتفاقی افتاده، چه کسانی یا چه نهادی درگیر هستند، و نتیجه یا نکته مهم خبر.
4. شرح می تواند 1 یا 2 جمله باشد.
5. کل شرح باید در یک خط نوشته شود.
6. حداکثر طول شرح {COMPRESSION_TARGET_MAX_WORDS} کلمه است؛ خودت در همان محدوده بمان و جمله را کامل تمام کن.
7. شرح نباید صرفاً بازنویسی تیتر باشد.
8. از توضیحات فرعی، نقل قول طولانی، و جزئیات غیرضروری خودداری کن.
9. اگر خروجی بیش از {COMPRESSION_TARGET_MAX_WORDS} کلمه شد، آن را دوباره فشرده کن.

قوانین سخت:
- هیچ متن اضافی قبل یا بعد از خروجی ننویس.
- فقط همین قالب را برگردان.
- تیتر و شرح نباید عملاً یک جمله یا یک بیان تکراری باشند.

قالب خروجی دقیق:
تیتر:
<headline>

شرح:
<one_line_summary>

ورودی:
تیتر خبر:
{article.get('original_title', '')}

متن خبر:
{article.get('content', '')[:2500]}
""".strip()

    def classify_article(
        self,
        article: Dict[str, Any],
        database: Dict[str, Any],
        costs: Dict[str, Any],
        logger: CycleLogger,
    ) -> Dict[str, Any]:
        evidence = build_classification_evidence(article)
        memory_text = classification_memory_prompt_text(database)
        result = self.call_llm_json(self.classification_prompt(article, evidence, memory_text), "classification", costs, logger)
        article["classification_category"] = normalize_category(result.get("category", evidence["suggested_category"]))
        article["classification_confidence"] = normalize_classification_confidence(result.get("confidence", "متوسط"))
        article["classification_rationale"] = normalize_text(result.get("reason", result.get("rationale", "")))
        article["classification_status"] = "success"
        article["classification_method"] = "llm_with_memory" if CLASSIFICATION_MEMORY_PROMPT_ENABLED else "llm_only"

        article["classification_version"] = CLASSIFICATION_VERSION
        article["classification_prompt_memory_version"] = CLASSIFICATION_PROMPT_MEMORY_VERSION
        fallback_memory_keywords = evidence["economics_hits"] + evidence["security_hits"]
        article["classification_memory_keywords"] = normalize_memory_keywords(
            result.get("memory_keywords", []),
            fallback_memory_keywords,
        )
        article["classification_memory_logic"] = normalize_memory_logic(
            result.get("memory_logic", ""),
            article["classification_rationale"],
        )
        article["classification_keyword_hits"] = {
            "economics": normalize_memory_keywords(result.get("matched_economics_keywords", []), evidence["economics_hits"]),
            "security": normalize_memory_keywords(result.get("matched_security_keywords", []), evidence["security_hits"]),
            "suggested_category": evidence["suggested_category"],
        }
        logger.event(
            "classification",
            "article_classified",
            url=article.get("url", ""),
            classified_as=article["classification_category"],
            confidence=article["classification_confidence"],
            method=article["classification_method"],
            suggested=evidence["suggested_category"],
            economics_hits=article["classification_keyword_hits"]["economics"],
            security_hits=article["classification_keyword_hits"]["security"],
            memory_keywords=article["classification_memory_keywords"],
            memory_logic=article["classification_memory_logic"],
        )
        return article

    def evaluate_article(
        self,
        article: Dict[str, Any],
        database: Dict[str, Any],
        costs: Dict[str, Any],
        logger: CycleLogger,
    ) -> Dict[str, Any]:
        # Check if evaluation should be skipped for 'other' category
        category = article.get('classification_category', 'other')
        if category == 'other':
            # Skip evaluation for 'other' category - cost savings
            article["confidence_occurrence"] = "نامشخص"
            article["gold_price_impact"] = "خیلی کم"
            article["security_relevance"] = "خیلی کم"
            article["gold_trend"] = "→"
            article["evaluation_rationale"] = "دسته other - ارزیابی نشد"
            article["evaluation_status"] = "skipped"
            article["evaluation_version"] = EVALUATION_VERSION
            return article
        
        # Use category-specific memory for better evaluation
        memory_text = category_specific_evaluation_memory(database, category)
        result = self.call_llm_json(self.evaluation_prompt(article, memory_text), "evaluation", costs, logger)
        
        # Parse results
        article["confidence_occurrence"] = validate_level(result.get("confidence_occurrence", "متوسط"))
        article["gold_trend"] = validate_trend(result.get("gold_trend", "?"))
        article["evaluation_rationale"] = normalize_text(result.get("rationale", ""))
        
        # Handle category-specific fields - set defaults for missing fields
        if category == "security":
            # Security evaluation returns security_relevance but not gold_price_impact
            article["security_relevance"] = validate_level(result.get("security_relevance", "متوسط"))
            article["gold_price_impact"] = "خیلی کم"
        elif category == "economics":
            # Economics evaluation returns gold_price_impact but not security_relevance
            article["gold_price_impact"] = validate_level(result.get("gold_price_impact", "متوسط"))
            article["security_relevance"] = "خیلی کم"
        else:  # security/economics
            # Mixed evaluation returns both
            article["gold_price_impact"] = validate_level(result.get("gold_price_impact", "متوسط"))
            article["security_relevance"] = validate_level(result.get("security_relevance", "متوسط"))
        
        article["evaluation_status"] = "success"
        article["evaluation_version"] = EVALUATION_VERSION
        return article

    def compress_article(
        self,
        article: Dict[str, Any],
        costs: Dict[str, Any],
        logger: CycleLogger,
    ) -> Dict[str, Any]:
        prompt = self.compression_prompt(article)
        for attempt in range(2):
            response = self.call_llm_text(prompt, "compression", costs, logger)
            headline_match = re.search(r"تیتر:\s*(.+?)\s*شرح:", response, re.DOTALL)
            summary_match = re.search(r"شرح:\s*(.+)", response, re.DOTALL)
            if not headline_match or not summary_match:
                if attempt == 1:
                    raise ValueError("Compression response format was invalid")
                prompt += "\n\nیادآوری سخت: تیتر و شرح نباید تقریبا یکسان باشند و شرح باید جزئیات مهم خبر را حمل کند."
                continue

            optimized_title = compact_sentence(headline_match.group(1))
            one_line_description = compact_sentence(summary_match.group(1))
            word_count = len(one_line_description.split())
            if word_count > COMPRESSION_TARGET_MAX_WORDS:
                logger.event(
                    "llm",
                    "compression_over_target_words",
                    url=article.get("url", ""),
                    words=word_count,
                    target=COMPRESSION_TARGET_MAX_WORDS,
                )

            if similarity_ratio(optimized_title, one_line_description) < 0.86:
                break

            logger.event(
                "llm",
                "compression_similarity_retry",
                url=article.get("url", ""),
                attempt=attempt + 1,
                title=optimized_title,
                summary=one_line_description,
            )
            if attempt == 1:
                raise ValueError("Compression response collapsed title and summary into near-duplicates")
            prompt += "\n\nیادآوری سخت: تیتر باید فقط از تیتر اصلی الهام بگیرد و شرح باید حتما محتوای خبر را با جزئیات بیشتر خلاصه کند."

        article["optimized_title"] = optimized_title
        article["one_line_description"] = one_line_description
        article["compression_status"] = "success"
        article["compression_version"] = COMPRESSION_VERSION
        return article

    def should_reprocess(self, article: Dict[str, Any]) -> bool:
        if article.get("duplicate_of"):
            return False
        if not article.get("processed"):
            return True
        if article.get("classification_version") and article.get("classification_status") not in ("success", "duplicate"):
            return True
        if article.get("classification_status") == "failed":
            return True
        if article.get("evaluation_status") == "failed":
            return True
        if article.get("compression_status") == "failed":
            return True
        if article.get("classification_version") != CLASSIFICATION_VERSION:
            return True
        if article.get("classification_prompt_memory_version") != CLASSIFICATION_PROMPT_MEMORY_VERSION:
            return True
        if article.get("evaluation_version") != EVALUATION_VERSION:
            return True
        if article.get("compression_version") != COMPRESSION_VERSION:
            return True
        return False

    def fetch_list_page_entries(self, page: int, logger: CycleLogger) -> List[Dict[str, Any]]:
        candidate_urls = [f"{BASE_LIST_URL}/?page={page}"]
        if page == 1:
            candidate_urls.append(BASE_LIST_URL)

        best_entries: List[Dict[str, Any]] = []
        best_page_url = candidate_urls[0]
        last_error: Optional[Exception] = None

        for attempt in range(1, LIST_PAGE_RETRY_COUNT + 1):
            for page_url in candidate_urls:
                try:
                    response = self.session.get(page_url, timeout=REQUEST_TIMEOUT)
                    response.raise_for_status()
                    page_entries = parse_listing_entries(response.text, logger)
                except Exception as exc:
                    last_error = exc
                    logger.warning(f"List page failed {page_url}: {exc}")
                    logger.event("errors", "list_page_failed", page=page, url=page_url, error=str(exc))
                    continue

                if len(page_entries) > len(best_entries):
                    best_entries = page_entries
                    best_page_url = page_url

                if len(page_entries) >= EXPECTED_LINKS_PER_PAGE:
                    if len(page_entries) > EXPECTED_LINKS_PER_PAGE:
                        logger.event(
                            "fetch",
                            "page_link_overflow",
                            page=page,
                            url=page_url,
                            found_links=len(page_entries),
                        )
                    return page_entries[:EXPECTED_LINKS_PER_PAGE]

            if attempt < LIST_PAGE_RETRY_COUNT:
                time.sleep(REQUEST_DELAY_SECONDS)

        if not best_entries and last_error:
            raise last_error

        logger.warning(
            f"Page {page} returned {len(best_entries)} links at {best_page_url}; expected {EXPECTED_LINKS_PER_PAGE}"
        )
        logger.event(
            "fetch",
            "page_link_shortfall",
            page=page,
            url=best_page_url,
            found_links=len(best_entries),
            expected_links=EXPECTED_LINKS_PER_PAGE,
        )
        return best_entries[:EXPECTED_LINKS_PER_PAGE]

    def fetch_cycle_articles(
        self,
        database: Dict[str, Any],
        cycle_id: str,
        logger: CycleLogger,
        stats: CycleStats,
    ) -> List[str]:
        window_values = {self.window["today"], self.window["yesterday"]}
        stale_pages = 0
        links_seen: set[str] = set()
        process_urls: List[str] = []
        hash_to_primary = {
            item.get("content_hash"): url
            for url, item in database.items()
            if item.get("content_hash") and not item.get("duplicate_of")
        }
        identity_to_primary = {
            article_identity_key(item): url
            for url, item in database.items()
            if not item.get("duplicate_of")
        }

        for page in range(1, HARD_MAX_PAGES + 1):
            stats.pages_scanned += 1
            try:
                page_entries = self.fetch_list_page_entries(page, logger)
            except Exception as exc:
                logger.error(f"List page failed for page {page}: {exc}")
                continue

            page_hits = 0
            print_stage("FETCH", f"Page {page}: {len(page_entries)} links")

            for entry in page_entries:
                url = entry["url"]
                entry_identity = entry.get("article_identity_key") or article_identity_key(entry)
                if entry_identity in links_seen:
                    continue
                links_seen.add(entry_identity)
                stats.links_seen += 1

                in_window = entry["listing_published_at_persian"] in window_values
                uncertain_allowed = entry["date_uncertain"] and page <= UNCERTAIN_DATE_EARLY_PAGE_LIMIT
                if not in_window and not uncertain_allowed:
                    stats.skipped_out_of_window += 1
                    logger.event(
                        "date_filter",
                        "skip_out_of_window",
                        url=url,
                        published_at_persian=entry["listing_published_at_persian"],
                        page=page,
                    )
                    continue

                page_hits += 1
                primary_storage_url = identity_to_primary.get(entry_identity)
                existing = database.get(primary_storage_url, {}) if primary_storage_url else database.get(url, {})
                if uncertain_allowed and entry["date_uncertain"]:
                    stats.uncertain_included += 1
                    logger.event("date_filter", "include_uncertain", url=url, page=page)

                needs_page_fetch = not existing or not existing.get("content")
                if existing and not needs_page_fetch and self.should_reprocess(existing):
                    article_data = {
                        "url": url,
                        "original_title": existing.get("original_title") or entry["listing_title"],
                        "lead": existing.get("lead", ""),
                        "content": existing.get("content", ""),
                        "source": existing.get("source", "خبر فوری"),
                        "content_hash": existing.get("content_hash", ""),
                        "published_at_gregorian": existing.get("published_at_gregorian") or entry["listing_published_at_gregorian"],
                        "published_at_persian": existing.get("published_at_persian") or entry["listing_published_at_persian"],
                        "published_time": existing.get("published_time") or entry["listing_published_time"],
                        "date_uncertain": existing.get("date_uncertain", entry["date_uncertain"]),
                        "fetched_at": now_tehran().isoformat(),
                    }
                elif existing and not needs_page_fetch and not self.should_reprocess(existing):
                    existing["last_seen_cycle"] = cycle_id
                    existing["listing_published_at_gregorian"] = entry["listing_published_at_gregorian"]
                    existing["listing_published_at_persian"] = entry["listing_published_at_persian"]
                    existing["listing_published_time"] = entry["listing_published_time"]
                    database[url] = existing
                    continue
                else:
                    article_data = extract_article_data(self.session, url, logger, entry)
                    if not article_data:
                        stats.failed += 1
                        continue

                article_identity = article_identity_key(article_data)
                storage_url = primary_storage_url or identity_to_primary.get(article_identity) or url
                existing = database.get(storage_url, existing)
                was_existing = storage_url in database
                primary_url = hash_to_primary.get(article_data["content_hash"])
                duplicate_of = None
                if primary_url and primary_url != storage_url:
                    duplicate_of = primary_url
                    stats.duplicates += 1

                merged = {
                    "url": url,
                    "url_identity_key": article_identity,
                    "source": article_data["source"],
                    "original_title": article_data["original_title"],
                    "optimized_title": existing.get("optimized_title", ""),
                    "one_line_description": existing.get("one_line_description", ""),
                    "lead": article_data["lead"],
                    "content": article_data["content"],
                    "content_hash": article_data["content_hash"],
                    "published_at_gregorian": article_data["published_at_gregorian"],
                    "published_at_persian": article_data["published_at_persian"],
                    "published_time": article_data["published_time"],
                    "listing_published_at_gregorian": entry["listing_published_at_gregorian"],
                    "listing_published_at_persian": entry["listing_published_at_persian"],
                    "listing_published_time": entry["listing_published_time"],
                    "fetched_at": article_data["fetched_at"],
                    "first_seen_cycle": existing.get("first_seen_cycle", cycle_id),
                    "last_seen_cycle": cycle_id,
                    "processed": existing.get("processed", False),
                    "classification_category": existing.get("classification_category", ""),
                    "classification_confidence": existing.get("classification_confidence", ""),
                    "classification_rationale": existing.get("classification_rationale", ""),
                    "classification_status": "duplicate" if duplicate_of else existing.get("classification_status", "pending"),
                    "classification_version": existing.get("classification_version"),
                    "classification_method": existing.get("classification_method", ""),
                    "classification_prompt_memory_version": existing.get("classification_prompt_memory_version", ""),
                    "classification_memory_keywords": existing.get("classification_memory_keywords", []),
                    "classification_memory_logic": existing.get("classification_memory_logic", ""),
                    "classification_keyword_hits": existing.get(
                        "classification_keyword_hits",
                        {"economics": [], "security": [], "suggested_category": "other"},
                    ),
                    "evaluation_status": "duplicate" if duplicate_of else existing.get("evaluation_status", "pending"),
                    "compression_status": "duplicate" if duplicate_of else existing.get("compression_status", "pending"),
                    "evaluation_version": existing.get("evaluation_version"),
                    "compression_version": existing.get("compression_version"),
                    "notified": False if duplicate_of else existing.get("notified", False),
                    "duplicate_of": duplicate_of,
                    "date_uncertain": article_data["date_uncertain"],
                    "confidence_occurrence": existing.get("confidence_occurrence", ""),
                    "gold_price_impact": existing.get("gold_price_impact", ""),
                    "security_relevance": existing.get("security_relevance", ""),
                    "gold_trend": existing.get("gold_trend", ""),
                    "notification_status": WORKBOOK_NOTIFICATION_FALSE,
                }
                database[storage_url] = merged

                if duplicate_of:
                    database[storage_url]["processed"] = True
                    continue

                hash_to_primary.setdefault(article_data["content_hash"], storage_url)
                identity_to_primary.setdefault(article_identity, storage_url)

                if not was_existing:
                    stats.new_articles += 1

                if self.should_reprocess(database[storage_url]):
                    process_urls.append(storage_url)

                if needs_page_fetch:
                    time.sleep(REQUEST_DELAY_SECONDS)

            stale_pages = stale_pages + 1 if page_hits == 0 else 0
            logger.event("fetch", "page_processed", page=page, page_hits=page_hits, stale_pages=stale_pages)
            if stale_pages >= STOP_AFTER_STALE_PAGES:
                print_stage("FETCH", f"Stopping after {stale_pages} stale pages")
                break

        return list(dict.fromkeys(process_urls))

    def persist_outputs(
        self,
        database: Dict[str, Any],
        tracking: Dict[str, Any],
        costs: Dict[str, Any],
        changed_dates: Optional[Set[str]] = None,
    ) -> None:
        """Write core JSON state + fast exports after each article (interrupt-safe)."""
        self.save_state(database, tracking, costs, rebuild_auxiliary=False)
        self.rebuild_notifications(database)
        self.rebuild_excel(database, target_dates=changed_dates)

    def process_articles(
        self,
        urls: Iterable[str],
        database: Dict[str, Any],
        tracking: Dict[str, Any],
        costs: Dict[str, Any],
        logger: CycleLogger,
        stats: CycleStats,
    ):
        for idx, url in enumerate(urls, start=1):
            article = database[url]
            print_stage("CLASSIFY", f"{idx}. {article.get('original_title', '')[:70]}")
            try:
                article = self.classify_article(article, database, costs, logger)
                stats.classified += 1
                category = article["classification_category"]
                if category == "security":
                    stats.security_count += 1
                elif category == "economics":
                    stats.economics_count += 1
                elif category == "security/economics":
                    stats.security_economics_count += 1
                else:
                    stats.other_count += 1

                print_stage(
                    "CLASSIFY",
                    f"→ {category} | security hits={len(article['classification_keyword_hits']['security'])} | "
                    f"economics hits={len(article['classification_keyword_hits']['economics'])}",
                )

                if category == "other":
                    article["processed"] = True
                    article["evaluation_status"] = "skipped_other"
                    article["compression_status"] = "skipped_other"
                    article["evaluation_version"] = EVALUATION_VERSION
                    article["compression_version"] = COMPRESSION_VERSION
                    article["optimized_title"] = ""
                    article["one_line_description"] = ""
                    article["notified"] = False
                    article["notification_status"] = WORKBOOK_NOTIFICATION_FALSE
                    database[url] = article
                    stats.skipped_other += 1
                    print_stage("SKIP", "Category is other, skipping evaluation and compression")
                    self.persist_outputs(database, tracking, costs, {article.get("published_at_persian", "")})
                    continue

                print_stage("LLM", f"Evaluating {category} article")
                article = self.evaluate_article(article, database, costs, logger)
                stats.evaluated += 1
                article = self.compress_article(article, costs, logger)
                stats.compressed += 1
                scores = {
                    "confidence": article["confidence_occurrence"],
                    "gold_impact": article["gold_price_impact"],
                    "security_relevance": article["security_relevance"],
                }
                article["notification_status"] = calculate_notification_status(scores)
                article["notified"] = article["notification_status"] == WORKBOOK_NOTIFICATION_TRUE
                article["processed"] = True
                if article["notified"]:
                    stats.notifying += 1
                database[url] = article
            except Exception as exc:
                article["processed"] = True
                article["evaluation_status"] = "failed"
                article["compression_status"] = "failed"
                article["notified"] = False
                article["notification_status"] = WORKBOOK_NOTIFICATION_FALSE
                database[url] = article
                stats.failed += 1
                logger.error(f"Processing failed for {url}: {exc}")
                logger.event("errors", "processing_failed", url=url, error=str(exc))

            self.persist_outputs(database, tracking, costs, {article.get("published_at_persian", "")})

    def rebuild_notifications(self, database: Dict[str, Any]):
        self.repair_notification_flags(database)
        articles = unique_articles_for_output(
            item for item in database.values() if item.get("notified") and not item.get("duplicate_of")
        )
        articles.sort(key=article_recency_sort_tuple, reverse=True)
        body = "\n".join(build_human_news_entry(article) for article in articles)
        IMPORTANT_NEWS_FILE.write_text(body, encoding="utf-8")
    
    def rebuild_txt_files(self, database: Dict[str, Any]):
        """Generate category-specific TXT files."""
        categories = {
            "security": TXT_DIR / "security_news.txt",
            "economics": TXT_DIR / "economics_news.txt",
            "security/economics": TXT_DIR / "security_economics_news.txt",
            "other": TXT_DIR / "other_news.txt",
        }
        
        # Organize articles by category
        articles_by_category = {cat: [] for cat in categories.keys()}
        
        for article in database.values():
            if article.get("duplicate_of"):
                continue
            if not article.get("processed"):
                continue
            
            category = article.get("classification_category", "other")
            if category in articles_by_category:
                articles_by_category[category].append(article)
        
        # Generate TXT files for each category
        for category, filepath in categories.items():
            articles = articles_by_category[category]
            articles.sort(key=article_recency_sort_tuple, reverse=True)
            
            lines = []
            for article in articles:
                title = normalize_text(article.get("optimized_title", article.get("original_title", "")))
                
                # For 'other' category, use lead instead of summary (no API cost)
                if category == "other":
                    lead = normalize_text(article.get("lead", ""))
                    # Wrap at 100 chars for better readability
                    words = lead.split()
                    wrapped_lines = []
                    current_line = []
                    current_length = 0
                    for word in words:
                        if current_length + len(word) + 1 > 100:
                            wrapped_lines.append(" ".join(current_line))
                            current_line = [word]
                            current_length = len(word)
                        else:
                            current_line.append(word)
                            current_length += len(word) + 1
                    if current_line:
                        wrapped_lines.append(" ".join(current_line))
                    summary = "\n".join(wrapped_lines) if wrapped_lines else lead
                else:
                    summary = normalize_text(article.get("one_line_description", ""))
                
                source = normalize_text(article.get("source", ""))
                date = persian_date_display_full(article.get("published_at_persian", ""))
                time = normalize_text(article.get("published_time", ""))
                
                # Add separator before article (except first one)
                if lines:
                    lines.append("")
                
                if title:
                    lines.append(title)
                    lines.append("")
                lines.append(summary)
                lines.append("")
                lines.append(f"{source} | {date} | {time}")
                lines.append("-" * 50)
            
            filepath.write_text("\n".join(lines) + "\n" if lines else "", encoding="utf-8")

    @staticmethod
    def repair_notification_flags(database: Dict[str, Any]) -> None:
        for url, article in database.items():
            if not article.get("processed") or article.get("duplicate_of") or article.get("classification_category") == "other":
                continue
            scores = {
                "confidence": article.get("confidence_occurrence", ""),
                "gold_impact": article.get("gold_price_impact", ""),
                "security_relevance": article.get("security_relevance", ""),
            }
            fresh_status = calculate_notification_status(scores)
            fresh_notified = fresh_status == WORKBOOK_NOTIFICATION_TRUE
            if article.get("notification_status") != fresh_status or article.get("notified") != fresh_notified:
                article["notification_status"] = fresh_status
                article["notified"] = fresh_notified

    def rebuild_excel(self, database: Dict[str, Any], target_dates: Optional[Set[str]] = None):
        self.repair_notification_flags(database)
        rows = []
        for article in unique_articles_for_output(database.values()):
            if article.get("duplicate_of"):
                continue
            if not article.get("processed"):
                continue
            if article.get("classification_category") == "other":
                continue
            rows.append(
                {
                    "_sort_published_at_gregorian": article.get("published_at_gregorian", ""),
                    "_published_at_persian": article.get("published_at_persian", ""),
                    "تاریخ انتشار": persian_date_display_full(article.get("published_at_persian", "")),
                    "ساعت انتشار": normalize_text(article.get("published_time", "")),
                    "منبع": normalize_text(article.get("source", "")),
                    "شناسه خبر": 0,
                    "تیتر خبر": normalize_text(article.get("optimized_title", "")),
                    "اطمینان از وقوع خبر": article.get("confidence_occurrence", ""),
                    "چقدر بر تغییر قیمت طلا اثر دارد؟": article.get("gold_price_impact", ""),
                    "چقدربا امنیت مرتبط است ؟": article.get("security_relevance", ""),
                    "جهت طلا": trend_display_value(article.get("gold_trend", "")),
                    "وضعیت اطلاع رسانی": article.get("notification_status", WORKBOOK_NOTIFICATION_FALSE),
                    "توضیحات": normalize_text(article.get("one_line_description", "")),
                    "لینک": normalize_text(article.get("url", "")),
                }
            )
        rows.sort(key=lambda item: (item.get("_sort_published_at_gregorian", ""), item.get("لینک", "")))
        for idx, item in enumerate(rows, start=1):
            item["شناسه خبر"] = idx

        rows_by_persian_date: Dict[str, List[Dict[str, Any]]] = {}
        for row in rows:
            published_date = row.get("_published_at_persian", "")
            if not published_date:
                continue
            rows_by_persian_date.setdefault(
                published_date,
                [],
            ).append({key: value for key, value in row.items() if not key.startswith("_")})

        if target_dates is not None:
            filtered_dates = {normalize_text(item) for item in target_dates if normalize_text(item)}
            for persian_date in filtered_dates:
                daily_path = workbook_daily_path(persian_date)
                rows_for_day = rows_by_persian_date.get(persian_date, [])
                if rows_for_day:
                    build_news_workbook(rows_for_day, daily_path)
                elif daily_path.exists():
                    daily_path.unlink()
            return

        for old_daily_file in EXCEL_DIR.glob(f"{EXCEL_BASE_NAME} - *.xlsx"):
            old_daily_file.unlink()
        if EXCEL_OUTPUT_FILE.exists():
            EXCEL_OUTPUT_FILE.unlink()

        for persian_date in sorted(rows_by_persian_date):
            build_news_workbook(rows_by_persian_date[persian_date], workbook_daily_path(persian_date))

    def run_cycle(self) -> CycleStats:
        ensure_runtime_layout()
        cycle_id = now_tehran().strftime("%Y%m%d_%H%M%S")
        stats = CycleStats(cycle_id=cycle_id)
        logger = CycleLogger(cycle_id)
        database, tracking, costs = self.load_state()
        costs.setdefault("cycle", {"calls": 0, "input_tokens": 0, "output_tokens": 0, "total_tokens": 0, "total_cost": 0.0, "stages": {}})

        tracking["last_started_cycle"] = cycle_id
        window_display = f"{self.window['today']} / {self.window['yesterday']}"
        print_box(
            "NEWS MONITOR",
            [
                ("Cycle", cycle_id),
                ("Window (Persian)", window_display),
                ("Model", API_MODEL),
                ("Runtime Folder", str(RUNTIME_DIR)),
            ],
        )

        logger.info(f"Cycle started with window {window_display}")
        logger.event("cycle", "start", window=self.window)
        costs["cycle"] = {
            "cycle_id": cycle_id,
            "calls": 0,
            "input_tokens": 0,
            "output_tokens": 0,
            "total_tokens": 0,
            "total_cost": 0.0,
            "stages": {},
            "started_at": now_tehran().isoformat(),
        }

        process_urls = self.fetch_cycle_articles(database, cycle_id, logger, stats)
        print_stage("FETCH", f"Process queue: {len(process_urls)} articles")
        self.process_articles(process_urls, database, tracking, costs, logger, stats)
        self.rebuild_notifications(database)
        self.rebuild_excel(database)

        costs["last_cycle"] = costs["cycle"] | {"finished_at": now_tehran().isoformat()}
        append_cost_run(costs, mode="live_cycle")
        tracking.setdefault("cycles", []).append(
            {
                "cycle_id": cycle_id,
                "started_at": costs["cycle"].get("started_at"),
                "finished_at": now_tehran().isoformat(),
                "window": self.window,
                "pages_scanned": stats.pages_scanned,
                "links_seen": stats.links_seen,
                "new_articles": stats.new_articles,
                "duplicates": stats.duplicates,
                "classified": stats.classified,
                "evaluated": stats.evaluated,
                "compressed": stats.compressed,
                "notifying": stats.notifying,
                "skipped_other": stats.skipped_other,
                "security_count": stats.security_count,
                "economics_count": stats.economics_count,
                "security_economics_count": stats.security_economics_count,
                "other_count": stats.other_count,
                "failed": stats.failed,
                "api_cost": costs["cycle"]["total_cost"],
            }
        )
        tracking["cycles"] = tracking["cycles"][-50:]
        tracking["last_successful_cycle"] = cycle_id
        self.save_state(database, tracking, costs)

        print_box(
            "CYCLE SUMMARY",
            [
                ("Pages scanned", str(stats.pages_scanned)),
                ("Links seen", str(stats.links_seen)),
                ("New articles", str(stats.new_articles)),
                ("Duplicates", str(stats.duplicates)),
                ("Classified", str(stats.classified)),
                ("Skipped other", str(stats.skipped_other)),
                ("Security", str(stats.security_count)),
                ("Economics", str(stats.economics_count)),
                ("Security/Economics", str(stats.security_economics_count)),
                ("Other", str(stats.other_count)),
                ("Evaluated", str(stats.evaluated)),
                ("Compressed", str(stats.compressed)),
                ("Notifying", str(stats.notifying)),
                ("Cycle API cost", f"${costs['cycle']['total_cost']:.6f}"),
            ],
        )
        return stats

    def show_stats(self):
        database, tracking, costs = self.load_state()
        notifying = sum(1 for item in database.values() if item.get("notified") and not item.get("duplicate_of"))
        processed = sum(1 for item in database.values() if item.get("processed") and not item.get("duplicate_of"))
        print_box(
            "PIPELINE STATS",
            [
                ("Tracked articles", str(len(database))),
                ("Processed articles", str(processed)),
                ("Notifying articles", str(notifying)),
                ("Last successful cycle", str(tracking.get("last_successful_cycle"))),
                ("Lifetime API calls", str(costs.get("lifetime", {}).get("calls", 0))),
                ("Lifetime API cost", f"${costs.get('lifetime', {}).get('total_cost', 0.0):.6f}"),
            ],
        )


# ==================== COMMANDS ====================
def reset_runtime():
    if RUNTIME_DIR.exists():
        shutil.rmtree(RUNTIME_DIR)
    ensure_runtime_layout()
    IMPORTANT_NEWS_FILE.write_text("", encoding="utf-8")
    CLASSIFICATION_MEMORY_FILE.write_text("# Classification Memory\n\n", encoding="utf-8")
    if EXCEL_OUTPUT_FILE.exists():
        EXCEL_OUTPUT_FILE.unlink()
    save_json(NEWS_DB_FILE, {})
    save_json(
        TRACKING_FILE,
        {
            "schema_version": SCHEMA_VERSION,
            "last_successful_cycle": None,
            "last_started_cycle": None,
            "cycles": [],
        },
    )
    save_json(API_COST_FILE, cost_state_for_save(default_cost_state()))
    save_json(PROMPT_MEMORY_PATH, {"updated_at": now_tehran().isoformat(), "examples": []})
    save_json(OTHER_NEWS_FILE, {"updated_at": now_tehran().isoformat(), "count": 0, "articles": []})


def run_single_cycle() -> CycleStats:
    pipeline = NewsPipeline()
    return pipeline.run_cycle()


def run_loop_mode(interval_minutes: int = LOOP_INTERVAL_MINUTES):
    pipeline = NewsPipeline()
    interval_minutes = max(1, interval_minutes)
    while True:
        try:
            pipeline.run_cycle()
        except KeyboardInterrupt:
            raise
        except Exception as exc:
            print_stage("ERROR", str(exc))
        print_stage("SLEEP", f"Waiting {interval_minutes} minutes for next cycle")
        time.sleep(interval_minutes * 60)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Cycle-based news monitoring pipeline")
    subparsers = parser.add_subparsers(dest="command", required=True)

    subparsers.add_parser("run-once", help="Run one fetch/evaluation cycle")

    loop_parser = subparsers.add_parser("run-loop", help="Run continuous cycles")
    loop_parser.add_argument("--interval-minutes", type=int, default=LOOP_INTERVAL_MINUTES)

    subparsers.add_parser("rebuild-notifications", help="Rebuild important_news.txt from canonical state")
    subparsers.add_parser("rebuild-excel", help="Rebuild Excel export from canonical state")
    subparsers.add_parser("stats", help="Show pipeline stats")
    subparsers.add_parser("reset-runtime", help="Delete runtime outputs and recreate clean layout")
    return parser.parse_args()


def main():
    args = parse_args()
    pipeline = NewsPipeline()

    if args.command == "reset-runtime":
        reset_runtime()
        print_stage("RESET", f"Runtime reset completed at {RUNTIME_DIR}")
        return

    if args.command == "stats":
        pipeline.show_stats()
        return

    if args.command == "rebuild-notifications":
        database, tracking, costs = pipeline.load_state()
        pipeline.rebuild_notifications(database)
        pipeline.save_state(database, tracking, costs)
        print_stage("EXPORT", f"Rebuilt {IMPORTANT_NEWS_FILE}")
        return

    if args.command == "rebuild-excel":
        database, tracking, costs = pipeline.load_state()
        pipeline.rebuild_excel(database)
        pipeline.save_state(database, tracking, costs)
        print_stage("EXPORT", f"Rebuilt daily Excel files in {EXCEL_DIR}")
        return

    if args.command == "run-once":
        run_single_cycle()
        return

    if args.command == "run-loop":
        run_loop_mode(args.interval_minutes)


if __name__ == "__main__":
    try:
        run_single_cycle()
    except KeyboardInterrupt:
        print_stage("STOP", "Interrupted by user")
