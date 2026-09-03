"""Everything downstream of a human label: the review queue, few-shot examples, the golden
set, and A/B comparison of two already-run variants.

The golden set is never hand-written - it is exported from approved review rows, so the
truth a provider is measured against is the same truth a person recorded in the dashboard,
and it grows as review happens.
"""

from __future__ import annotations

import json
import sqlite3
from collections import defaultdict, deque
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from . import dag, text
from .prompts import ReviewedExample
from .providers import Provider
from .scoring import AXES, LEVELS, decide, level_score
from .sources import RawArticle

QUEUE_COLUMNS = [
    "review_id", "stratum", "status", "url", "title", "lead", "content", "source",
    "outlet", "reviewed_category", "confidence_occurrence", "gold_price_impact",
    "security_relevance", "gold_trend", "one_line", "reviewer_notes",
]


# ------------------------------------------------------------------------------ queue


def create_queue(conn: sqlite3.Connection, *, size: int = 100) -> int:
    """Create a deterministic, stratified queue without overwriting completed review.

    Strata are the cases worth a human's time: articles two prompts disagreed about, the
    `other` bucket, unevaluated ones, then each category round-robin.
    """
    rows = conn.execute("""
        SELECT a.id, c.category, counts.category_count, e.confidence_occurrence
        FROM articles a
        JOIN classifications c ON c.article_id=a.id
        JOIN (SELECT article_id, COUNT(DISTINCT category) AS category_count
              FROM classifications GROUP BY article_id) counts ON counts.article_id=a.id
        LEFT JOIN latest_evaluation e ON e.article_id=a.id
        WHERE a.duplicate_of IS NULL
        ORDER BY a.id
    """).fetchall()

    groups: dict[str, deque[sqlite3.Row]] = defaultdict(deque)
    seen: set[int] = set()
    for row in rows:
        if row["id"] in seen:
            continue
        seen.add(row["id"])
        if row["category_count"] > 1:
            groups["prompt_disagreement"].append(row)
        elif row["category"] == "other":
            groups["other"].append(row)
        elif row["confidence_occurrence"] is None:
            groups["unevaluated"].append(row)
        else:
            groups[row["category"]].append(row)

    chosen: list[tuple[sqlite3.Row, str]] = []
    labels = sorted(groups)
    while len(chosen) < size and labels:
        for label in list(labels):
            if groups[label]:
                chosen.append((groups[label].popleft(), label))
                if len(chosen) == size:
                    break
            if not groups[label]:
                labels.remove(label)
    for row, stratum in chosen:
        conn.execute(
            "INSERT OR IGNORE INTO review_cases(article_id,stratum,created_at) VALUES(?,?,?)",
            (row["id"], stratum, dag.utc_now()),
        )
    return len(chosen)


def _queue_rows(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute("""
        SELECT r.id AS review_id, r.stratum, r.status, a.url, a.original_title AS title,
               a.lead, a.content, a.source, a.original_outlet AS outlet, r.reviewed_category,
               r.confidence_occurrence, r.gold_price_impact, r.security_relevance,
               r.gold_trend, r.one_line, r.reviewer_notes
        FROM review_cases r JOIN articles a ON a.id=r.article_id
        ORDER BY r.stratum, r.id
    """).fetchall()


def export_queue(conn: sqlite3.Connection, path: Path) -> Path:
    """Write the queue as .xlsx (or .json) for offline labelling."""
    rows = _queue_rows(conn)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix == ".json":
        payload = [{key: row[key] for key in QUEUE_COLUMNS} for row in rows]
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return path
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review Queue"
    sheet.append(QUEUE_COLUMNS)
    for row in rows:
        sheet.append([row[key] for key in QUEUE_COLUMNS])
    sheet.freeze_panes = "A2"
    for column, width in (("E", 50), ("G", 80), ("P", 50)):
        sheet.column_dimensions[column].width = width
    workbook.save(path)
    return path


def import_queue(conn: sqlite3.Connection, path: Path) -> int:
    """Apply only rows explicitly marked approved; blank or pending rows stay untouched."""
    sheet = load_workbook(path, read_only=True, data_only=True).active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    updated = 0
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if str(row.get("status", "")).strip().lower() != "approved":
            continue
        updated += conn.execute(
            "UPDATE review_cases SET status='approved', reviewed_category=?,"
            " confidence_occurrence=?, gold_price_impact=?, security_relevance=?, gold_trend=?,"
            " one_line=?, reviewer_notes=?, reviewed_at=? WHERE id=? AND status='pending'",
            (row.get("reviewed_category"), row.get("confidence_occurrence"),
             row.get("gold_price_impact"), row.get("security_relevance"), row.get("gold_trend"),
             row.get("one_line"), row.get("reviewer_notes"), dag.utc_now(), row.get("review_id")),
        ).rowcount
    return updated


def reviewed_examples(
    conn: sqlite3.Connection, article: RawArticle, *, task: str,
    category: str | None = None, limit: int = 3,
) -> list[ReviewedExample]:
    """Approved labels closest to this article by title similarity, as few-shot examples."""
    rows = conn.execute("""
        SELECT a.original_title, r.reviewed_category, r.confidence_occurrence,
               r.gold_price_impact, r.security_relevance, r.gold_trend, r.one_line
        FROM review_cases r JOIN articles a ON a.id=r.article_id
        WHERE r.status='approved' AND r.reviewed_category IS NOT NULL
    """).fetchall()
    terms = text.trigrams(article.title)
    scored = []
    for row in rows:
        if category and row["reviewed_category"] not in {category, "security/economics"}:
            continue
        if task == "classify":
            output = {"category": row["reviewed_category"]}
        elif task == "evaluate":
            output = {axis: row[axis] for axis in (*AXES, "gold_trend")}
        else:
            output = {"one_line": row["one_line"]}
        scored.append((
            text.jaccard(terms, text.trigrams(row["original_title"])),
            ReviewedExample(row["reviewed_category"], row["original_title"],
                            json.dumps(output, ensure_ascii=False)),
        ))
    return [example for _, example in sorted(scored, key=lambda pair: pair[0], reverse=True)[:limit]]


# ------------------------------------------------------------------------- golden set


@dataclass(frozen=True)
class EvaluationCase:
    article: RawArticle
    category: str
    scores: dict[str, str | None]


def load_cases(path: Path) -> list[EvaluationCase]:
    return [
        EvaluationCase(RawArticle(**item["article"]), item["category"], item.get("scores", {}))
        for item in json.loads(path.read_text(encoding="utf-8"))
    ]


def build_golden(conn: sqlite3.Connection, path: Path) -> int:
    """Export every approved review as an evaluation case. Returns the case count.

    Only approved rows are eligible - a golden set seeded with the model's own answers
    would measure the model against itself and report perfect agreement. Axes the reviewer
    left unassessed are omitted rather than written as a level.
    """
    rows = conn.execute("""
        SELECT a.url, a.source, a.original_title, a.lead, a.content, a.original_outlet,
               a.published_at_gregorian, r.reviewed_category, r.confidence_occurrence,
               r.gold_price_impact, r.security_relevance
        FROM review_cases r JOIN articles a ON a.id = r.article_id
        WHERE r.status = 'approved' AND r.reviewed_category IS NOT NULL
        ORDER BY r.id
    """).fetchall()
    cases = [
        {
            "article": {
                "source": row["source"], "url": row["url"], "title": row["original_title"],
                "lead": row["lead"] or "", "content": row["content"] or "",
                "original_outlet": row["original_outlet"] or "",
                "published_at": row["published_at_gregorian"],
            },
            "category": row["reviewed_category"],
            "scores": {axis: row[axis] for axis in AXES if row[axis]},
        }
        for row in rows
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cases, ensure_ascii=False, indent=2), encoding="utf-8")
    return len(cases)


def weighted_kappa(expected: list[int], actual: list[int]) -> float | None:
    """Quadratic weighted kappa without a dependency; None when too few labels to compare."""
    if len(expected) < 2:
        return None
    size = len(LEVELS)
    observed = [[0] * size for _ in range(size)]
    for left, right in zip(expected, actual):
        observed[left - 1][right - 1] += 1
    row_totals = [sum(row) for row in observed]
    col_totals = [sum(observed[r][c] for r in range(size)) for c in range(size)]
    weight = lambda i, j: ((i - j) / (size - 1)) ** 2
    pairs = [(i, j) for i in range(size) for j in range(size)]
    total = len(expected)
    numerator = sum(weight(i, j) * observed[i][j] for i, j in pairs) / total
    denominator = sum(weight(i, j) * row_totals[i] * col_totals[j] for i, j in pairs) / total**2
    return 1.0 - numerator / denominator if denominator else 1.0


def evaluate(cases: list[EvaluationCase], provider: Provider) -> dict[str, Any]:
    """Score a provider against a golden set: category accuracy plus per-axis kappa."""
    categories = [provider.classify(case.article).data.category for case in cases]
    # One evaluate() call per case - it returns all three axes together, so calling it per
    # axis would triple the provider requests for the same answers.
    evaluations = [provider.evaluate(case.article, case.category).data for case in cases]
    kappa = {}
    for axis in AXES:
        expected, actual = [], []
        for case, evaluation in zip(cases, evaluations):
            wanted, got = case.scores.get(axis), getattr(evaluation, axis)
            if wanted in LEVELS and got in LEVELS:
                expected.append(level_score(wanted))
                actual.append(level_score(got))
        kappa[axis] = weighted_kappa(expected, actual)
    correct = sum(actual == case.category for actual, case in zip(categories, cases))
    return {
        "cases": len(cases),
        "category_accuracy": correct / len(cases) if cases else None,
        "kappa": kappa,
    }


# --------------------------------------------------------------------------- A/B diff


@dataclass(frozen=True)
class Variant:
    """One (provider, model, prompt_version) combination already run via `cli run`.
    `model=None` matches any model recorded for that provider."""

    provider: str
    model: str | None
    prompt_version: str


def _variant_rows(conn: sqlite3.Connection, variant: Variant) -> dict[int, dict[str, Any]]:
    params: list[Any] = [variant.provider, variant.prompt_version]
    model_clause = ""
    if variant.model:
        model_clause = "AND c.model = ?"
        params.append(variant.model)
    rows = conn.execute(
        f"""
        SELECT a.id AS article_id, a.original_title AS title, a.url AS url,
               c.category, c.confidence, c.rationale AS classify_rationale,
               e.confidence_occurrence, e.gold_price_impact, e.security_relevance,
               e.gold_trend, e.rationale AS eval_rationale
        FROM articles a
        JOIN classifications c ON c.article_id = a.id
            AND c.provider = ? AND c.prompt_version = ? {model_clause}
        LEFT JOIN evaluations e ON e.article_id = a.id AND e.provider = c.provider
            AND e.model = c.model AND e.prompt_version = c.prompt_version
        WHERE a.duplicate_of IS NULL
        """,
        params,
    ).fetchall()
    return {row["article_id"]: dict(row) for row in rows}


def diff_variants(conn: sqlite3.Connection, a: Variant, b: Variant) -> list[dict[str, Any]]:
    """Per-article diff between two already-run variants, over their shared articles.

    Inference is append-only, so running the same articles under two prompt versions just
    means two prior `run` invocations; this reads both instead of paying for a third call.
    Backs both `cli compare` (writes txt) and the dashboard's /compare page (renders HTML).
    """
    a_rows, b_rows = _variant_rows(conn, a), _variant_rows(conn, b)
    for label, rows, variant in (("A", a_rows, a), ("B", b_rows, b)):
        if not rows:
            raise ValueError(
                f"no rows for variant {label} ({variant.provider}/{variant.prompt_version});"
                " run it first"
            )

    def verdict(row: dict[str, Any]) -> tuple[str, str]:
        scores = (row["confidence_occurrence"], row["gold_price_impact"], row["security_relevance"])
        return row["category"], decide(*scores).status

    return [
        {
            "article_id": article_id,
            "title": a_rows[article_id]["title"],
            "url": a_rows[article_id]["url"],
            "agree": verdict(a_rows[article_id]) == verdict(b_rows[article_id]),
            "a": a_rows[article_id],
            "b": b_rows[article_id],
        }
        for article_id in sorted(set(a_rows) & set(b_rows))
    ]


def compare(conn: sqlite3.Connection, *, a: Variant, b: Variant, out_dir: Path) -> dict[str, Any]:
    """Write comparison_{same,different,all}.txt so a human can judge which variant is
    better before either becomes the default in config/routing.yaml."""
    def block(label: str, row: dict[str, Any]) -> str:
        return (
            f"[{label}] category={row['category']} confidence={row['confidence']}\n"
            f"  occurrence={row['confidence_occurrence']} gold_impact={row['gold_price_impact']}"
            f" security={row['security_relevance']} trend={row['gold_trend']}\n"
            f"  rationale: {row['eval_rationale'] or row['classify_rationale'] or ''}"
        )

    records = diff_variants(conn, a, b)
    buckets: dict[str, list[str]] = {"same": [], "different": [], "all": []}
    for record in records:
        rendered = (f"{record['title']}\n{record['url']}\n"
                    f"{block('A', record['a'])}\n{block('B', record['b'])}\n" + "-" * 50)
        verdict = "same" if record["agree"] else "different"
        buckets[verdict].append(rendered)
        buckets["all"].append(f"{verdict} | {rendered}")

    out_dir.mkdir(parents=True, exist_ok=True)
    for name, lines in buckets.items():
        (out_dir / f"comparison_{name}.txt").write_text("\n\n".join(lines), encoding="utf-8")
    return {
        "shared_articles": len(records),
        "same": len(buckets["same"]),
        "different": len(buckets["different"]),
        "out_dir": str(out_dir),
    }
