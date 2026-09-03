"""Daily Excel workbook and text feeds, built on the team's own template."""

from __future__ import annotations

import re
import shutil
import sqlite3
import warnings
import zipfile
from copy import copy
from datetime import time
from functools import lru_cache
from pathlib import Path

import jdatetime
from openpyxl import load_workbook

from . import config
from .prompts import CATEGORIES, GOLD_TRENDS
from .scoring import FLOOR, HIGH_BAR, HIGH_COUNT_REQUIRED, LEVELS, NOTIFY, NO_NOTIFY, decide

HEADERS = [
    "شناسه خبر", "تاریخ انتشار", "ساعت انتشار", "منبع", "تیتر خبر",
    "اطمینان از وقوع خبر", "چقدر بر تغییر قیمت طلا اثر دارد؟",
    "چقدربا امنیت مرتبط است ؟", "جهت طلا", "وضعیت اطلاع رسانی", "توضیحات", "لینک",
]
SHEET = "بررسی خبر"
MAX_STYLED_ROW = 504  # the template pre-styles to here; dropdowns should cover the same span
FIRST_DATA_ROW = 3


def _persian_date(value: str | None) -> str:
    if not value:
        return ""
    try:
        year, month, day = map(int, value.replace("/", "-").split("-"))
        return f"{day} {jdatetime.date.j_months_fa[month - 1]} {year}"
    except (ValueError, IndexError):
        return value


def workbook_filename(day: str) -> str:
    """The naming convention the team already files by - an ISO date would land the same
    content under names nobody recognises."""
    return f"ثبت و تحلیل خبر - {_persian_date(day) or day}.xlsx"


def _time(value: str | None) -> time | str:
    if not value:
        return ""
    try:
        hour, minute = map(int, value.split(":")[:2])
        return time(hour, minute)
    except ValueError:
        return value


def rows(conn: sqlite3.Connection) -> list[dict[str, str]]:
    """One record per canonical article, keyed by the workbook's own column headers."""
    query = """
    SELECT a.*, c.category, e.confidence_occurrence, e.gold_price_impact,
           e.security_relevance, e.gold_trend, s.optimized_title, s.one_line
    FROM articles a
    LEFT JOIN latest_classification c ON c.article_id=a.id
    LEFT JOIN latest_evaluation e ON e.article_id=a.id
    LEFT JOIN latest_summary s ON s.article_id=a.id
    WHERE a.duplicate_of IS NULL
    ORDER BY COALESCE(a.published_at_gregorian, a.fetched_at), a.url
    """
    return [
        {
            "_day": row["published_at_persian"] or "",
            "_category": row["category"] or "other",
            "شناسه خبر": str(index),
            "تاریخ انتشار": _persian_date(row["published_at_persian"]),
            "ساعت انتشار": row["published_time"] or "",
            "منبع": row["original_outlet"] or row["source"],
            "تیتر خبر": row["optimized_title"] or row["original_title"],
            "اطمینان از وقوع خبر": row["confidence_occurrence"] or "",
            "چقدر بر تغییر قیمت طلا اثر دارد؟": row["gold_price_impact"] or "",
            "چقدربا امنیت مرتبط است ؟": row["security_relevance"] or "",
            "جهت طلا": row["gold_trend"] or "",
            "وضعیت اطلاع رسانی": decide(
                row["confidence_occurrence"], row["gold_price_impact"], row["security_relevance"]
            ).status,
            "توضیحات": row["one_line"] or "",
            "لینک": row["url"],
        }
        for index, row in enumerate(conn.execute(query), 1)
    ]


def _formula(row: int) -> str:
    """The in-sheet notify formula, built from scoring.py's own thresholds and vocabulary
    so the workbook cannot drift into voting differently from decide()."""
    axes = f"F{row}:H{row}"
    high = "+".join(f'COUNTIF({axes},"{level}")' for level in LEVELS[HIGH_BAR - 1:])
    below_floor = "+".join(f'COUNTIF({axes},"{level}")' for level in LEVELS[: FLOOR - 1])
    return (f'=IF(AND({high}>={HIGH_COUNT_REQUIRED},{below_floor}=0),'
            f'"{NOTIFY}","{NO_NOTIFY}")')


@lru_cache(maxsize=1)
def _template_sheet_xml() -> bytes:
    """The template's sheet1.xml, read once - export_all() rebuilds a workbook per day."""
    with zipfile.ZipFile(config.WORKBOOK_TEMPLATE_PATH) as source:
        return source.read("xl/worksheets/sheet1.xml")


def _extract_ext_list(xml: bytes) -> tuple[int, int]:
    """(start, end) byte offsets of the <extLst> block, or (-1, -1) when absent."""
    start = xml.find(b"<extLst")
    end = xml.find(b"</extLst>", start)
    return (start, end + len(b"</extLst>")) if start >= 0 and end >= 0 else (-1, -1)


def _restore_template_extensions(path: Path) -> None:
    """Re-attach the template's extLst block, which openpyxl drops on save (it carries the
    conditional formatting and data validation extensions Excel expects)."""
    template_xml = _template_sheet_xml()
    ext_start, ext_end = _extract_ext_list(template_xml)
    if ext_start < 0:
        return
    extension = template_xml[ext_start:ext_end]

    with zipfile.ZipFile(path) as source:
        payload = {entry.filename: source.read(entry.filename) for entry in source.infolist()}
    xml = payload["xl/worksheets/sheet1.xml"]

    # openpyxl may omit namespaces the extension block references; copy any it dropped.
    def root_tag(document: bytes) -> tuple[int, int]:
        start = document.find(b"<worksheet")
        return start, document.find(b">", start)

    root_end = root_tag(xml)[1]
    template_start, template_end = root_tag(template_xml)
    namespaces = re.findall(
        rb'\s(xmlns(?::[A-Za-z0-9]+)?="[^"]+")', template_xml[template_start:template_end + 1]
    )
    missing = [ns for ns in namespaces if ns.split(b"=", 1)[0] not in xml[:root_end]]
    if missing:
        xml = xml[:root_end] + b" " + b" ".join(missing) + xml[root_end:]

    start, end = _extract_ext_list(xml)
    if start >= 0:
        xml = xml[:start] + xml[end:]
    payload["xl/worksheets/sheet1.xml"] = xml.replace(b"</worksheet>", extension + b"</worksheet>")

    temporary = path.with_suffix(".tmp")
    with zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as target:
        for name, content in payload.items():
            target.writestr(name, content)
    temporary.replace(path)


def _fix_validations(sheet) -> None:
    """Point the template's dropdowns at the values this pipeline actually writes.

    The template carries two stale ones: the score columns get a yes/no list from row 304
    down, and the gold-trend list stops at row 303 - so an analyst clicking a dropdown past
    row 303 is offered answers that do not belong in that column.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    sheet.data_validations.dataValidation = []
    for values, span in ((LEVELS, f"F3:H{MAX_STYLED_ROW}"), (GOLD_TRENDS, f"I3:I{MAX_STYLED_ROW}")):
        validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(span)


def _copy_row_style(sheet, source: int, target: int) -> None:
    for column in range(1, len(HEADERS) + 1):
        origin, destination = sheet.cell(source, column), sheet.cell(target, column)
        destination._style = copy(origin._style)
        for attribute in ("number_format", "protection", "alignment", "fill", "font", "border"):
            setattr(destination, attribute, copy(getattr(origin, attribute)))
    sheet.row_dimensions[target].height = sheet.row_dimensions[source].height


def build_workbook(records: list[dict[str, str]], target: Path) -> Path:
    """Clone the template into a daily workbook, preserving its styling.

    The template is the analyst's own source file - four sheets, including reference
    material. All 40 workbooks the team actually produced carry a single sheet, so the
    auxiliary ones are dropped here rather than shipped every day.
    """
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(config.WORKBOOK_TEMPLATE_PATH, target)
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message=r".*extension is not supported.*", category=UserWarning)
        workbook = load_workbook(target)
    for name in list(workbook.sheetnames):
        if name != SHEET:
            del workbook[name]
    sheet = workbook[SHEET]
    # The template's own header row is missing this cell, so every workbook the legacy
    # exporter produced from it had an unlabelled link column.
    sheet.cell(2, len(HEADERS)).value = HEADERS[-1]
    _fix_validations(sheet)

    last_existing = max(sheet.max_row, FIRST_DATA_ROW)
    for row in range(FIRST_DATA_ROW, last_existing + 1):
        for column in range(1, len(HEADERS) + 1):
            cell = sheet.cell(row, column)
            cell.value, cell.hyperlink = None, None
    for index, record in enumerate(records, FIRST_DATA_ROW):
        if index > last_existing:
            _copy_row_style(sheet, FIRST_DATA_ROW, index)
        for column, header in enumerate(HEADERS, 1):
            cell = sheet.cell(index, column)
            if header == "وضعیت اطلاع رسانی":
                cell.value = _formula(index)
            elif header == "ساعت انتشار":
                cell.value = _time(record[header])
            else:
                cell.value = record[header]
                if header == "لینک" and record[header]:
                    cell.hyperlink = record[header]
    workbook.save(target)
    workbook.close()
    _restore_template_extensions(target)
    return target


def _feed(records: list[dict[str, str]]) -> str:
    return "\n\n".join(
        f"{record['تیتر خبر']}\n\n{record['توضیحات']}\n\n"
        f"{record['منبع']} | {record['تاریخ انتشار']} | {record['ساعت انتشار']}\n" + "-" * 50
        for record in records
    ) + ("\n" if records else "")


def export_all(conn: sqlite3.Connection, directory: Path) -> dict[str, Path]:
    """One workbook per Jalali day, plus the notify feed and one text file per category."""
    directory.mkdir(parents=True, exist_ok=True)
    records = rows(conn)
    eligible = [record for record in records if record["_category"] != "other"]

    by_day: dict[str, list[dict[str, str]]] = {}
    for record in eligible:
        if record["_day"]:
            by_day.setdefault(record["_day"], []).append(record)

    files = {
        f"excel:{day}": build_workbook(day_rows, directory / "Excel Files" / workbook_filename(day))
        for day, day_rows in by_day.items()
    }
    files["important"] = directory / "important_news.txt"
    files["important"].write_text(
        _feed([r for r in eligible if r["وضعیت اطلاع رسانی"] == NOTIFY]), encoding="utf-8"
    )
    for category in CATEGORIES:
        path = directory / "TXT Files" / f"{category.replace('/', '_')}_news.txt"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(_feed([r for r in records if r["_category"] == category]), encoding="utf-8")
        files[f"text:{category}"] = path
    return files
