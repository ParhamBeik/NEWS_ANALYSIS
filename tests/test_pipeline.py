"""Ingest, deduplication, per-node routing, window backfill, export, and the CLI.

Integration-level: real SQLite, real dedup, real workbook writing - only the network and
the LLM are substituted, so the wiring between stages is what is under test.
"""

import argparse
import zipfile
from datetime import datetime, timedelta, timezone

import jdatetime
import pytest

from news_intel import cli, config, dag, db, dedupe, exports, pipeline, prompts, providers, sources
from news_intel.providers import FallbackProvider, RuleProvider
from news_intel.scoring import LEVELS
from news_intel.sources import RawArticle, SourceSpec

from test_providers import Failing, MeteredProvider  # noqa: E402 - sibling test module


def article(url="https://example.test/1", title="حمله موشکی به تاسیسات",
            *, source="khabarfoori", content="جزئیات خبر",
            published="2026-08-16T10:00:00+03:30", lead="خبر امنیتی اقتصادی"):
    return RawArticle(source=source, url=url, title=title, lead=lead, content=content,
                      original_outlet="ایسنا", published_at=published)


def register_sources(conn):
    for name, priority in (("khabarfoori", 1), ("mehr", 2), ("shahrekhabar", 3)):
        conn.execute(
            "INSERT OR REPLACE INTO sources(name,tier,priority,enabled) VALUES(?,2,?,1)",
            (name, priority),
        )


def add(conn, art, run_id="r1"):
    return pipeline.upsert_article(conn, art, run_id)


def canonical_count(conn):
    return conn.execute(
        "SELECT COUNT(*) c FROM articles WHERE duplicate_of IS NULL").fetchone()["c"]


def duplicate_map(conn):
    return {r["id"]: r["duplicate_of"] for r in conn.execute("SELECT id, duplicate_of FROM articles")}


class Recording(RuleProvider):
    """A rule provider that remembers which nodes called it."""

    def __init__(self, name, model):
        super().__init__(name=name, model=model)
        self.calls = []

    def classify(self, article, examples=()):
        self.calls.append("classify")
        return super().classify(article, examples)

    def evaluate(self, article, category, examples=()):
        self.calls.append("evaluate")
        return super().evaluate(article, category, examples)

    def summarize(self, article, examples=()):
        self.calls.append("summarize")
        return super().summarize(article, examples)


# ------------------------------------------------------------------------ processing


def test_process_is_idempotent_and_retains_unassessed_axes(conn):
    provider = RuleProvider()
    first = pipeline.process(conn, [article()], provider, run_id="run1")
    second = pipeline.process(conn, [article()], provider, run_id="run2")
    assert first == {"fetched": 1, "new": 1, "rejected": 0, "duplicate": 0,
                     "classified": 1, "evaluated": 1, "summarized": 1}
    assert second == {"fetched": 1, "new": 0, "rejected": 0, "duplicate": 0,
                      "classified": 0, "evaluated": 0, "summarized": 0}
    evaluation = conn.execute("SELECT * FROM evaluations").fetchone()
    assert evaluation["gold_price_impact"] == "زیاد"
    assert evaluation["security_relevance"] == "زیاد"


def test_a_rejected_article_is_quarantined_not_classified(conn):
    register_sources(conn)
    bad = RawArticle(source="khabarfoori", url="https://a/1", title="", lead="", content="")
    stats = pipeline.process(conn, [bad, article(url="https://a/2")], RuleProvider())
    assert stats["rejected"] == 1
    assert stats["classified"] == 1, "only the good article costs an inference call"
    rows = conn.execute("SELECT node, error_class FROM dead_letters").fetchall()
    assert [(r["node"], r["error_class"]) for r in rows] == [("quality", "missing_title")]


def test_provider_usage_is_persisted_in_node_events(conn):
    pipeline.process(conn, [article()], MeteredProvider(), run_id="metered")
    events = conn.execute(
        "SELECT tokens_in,tokens_out,cost_usd FROM node_events WHERE run_id='metered'").fetchall()
    assert len(events) == 3
    assert sum(row["tokens_in"] for row in events) == 30
    assert sum(row["tokens_out"] for row in events) == 15
    assert sum(row["cost_usd"] for row in events) == pytest.approx(0.03)


def test_an_incomplete_routing_map_is_refused_before_any_call(conn):
    with pytest.raises(config.ConfigError, match="evaluate"):
        pipeline.process(conn, [], {"classify": RuleProvider(), "summarize": RuleProvider()})


# --------------------------------------------------------------------- node routing


def test_each_node_records_the_model_that_actually_answered_it(conn):
    """A split run has to stay legible afterwards: rows carry their own model."""
    cheap, strong = Recording("rule", "small"), Recording("rule", "large")
    stats = pipeline.process(conn, [article(title="حمله موشکی و جهش قیمت طلا در بازار تهران")],
                             {"classify": cheap, "evaluate": strong, "summarize": cheap})
    assert stats["classified"] == 1 and stats["evaluated"] == 1
    assert cheap.calls == ["classify", "summarize"] and strong.calls == ["evaluate"]
    for table, model in (("classifications", "small"), ("evaluations", "large"), ("summaries", "small")):
        assert conn.execute(f"SELECT model FROM {table}").fetchone()["model"] == model


def test_a_single_provider_still_serves_every_node(conn):
    provider = Recording("rule", "solo")
    pipeline.process(conn, [article(title="حمله موشکی و جهش قیمت طلا در بازار تهران")], provider)
    assert sorted(provider.calls) == ["classify", "evaluate", "summarize"]


def test_changing_one_nodes_model_reruns_only_that_node(conn):
    """The existence check is keyed on provider+model, so a swap is not a full re-run."""
    item = article(title="حمله موشکی و جهش قیمت طلا در بازار تهران")
    first = {node: RuleProvider(name="rule", model="v1") for node in pipeline.NODES}
    pipeline.process(conn, [item], first)

    swapped = {**first, "evaluate": RuleProvider(name="rule", model="v2")}
    stats = pipeline.process(conn, [item], swapped)
    assert stats["classified"] == 0, "classify was unchanged and must not be paid for again"
    assert stats["evaluated"] == 1, "evaluate moved to a new model and must re-run"
    assert [r["model"] for r in conn.execute("SELECT model FROM evaluations ORDER BY id")] == ["v1", "v2"]


def test_a_fallback_route_records_the_backend_that_answered_and_reruns_only_once(conn):
    """`_already_ran()` used to key on FallbackProvider's composite name (e.g.
    "failing+fake"), which never matches the backend recorded on a row - so a route with a
    fallback re-classified, and re-billed, every article on every run."""
    wrapped = FallbackProvider(Failing(dag.Transient("down")), MeteredProvider())
    item = article(url="https://test/3", title="حمله موشکی به تاسیسات نفتی کشور")
    pipeline.process(conn, [item], wrapped, run_id="fallback-1")
    pipeline.process(conn, [item], wrapped, run_id="fallback-2")

    rows = conn.execute("SELECT provider, model FROM classifications").fetchall()
    assert len(rows) == 1, "the second run must recognise the work as already done"
    assert (rows[0]["provider"], rows[0]["model"]) == ("fake", "fake-v1")


# ------------------------------------------------------------------------ dedupe

class TestExactAndNearDuplicates:
    def test_identical_content_is_deduplicated(self, conn):
        register_sources(conn)
        add(conn, article("https://a.test/1"))
        assert add(conn, article("https://b.test/2"))[2] is True
        assert canonical_count(conn) == 1

    def test_encoding_variants_collide(self, conn):
        """Arabic KAF vs Persian KEHEH is the same story, not two."""
        register_sources(conn)
        add(conn, article("https://a.test/1", "حمله موشكی به تاسیسات"))  # Arabic KAF
        assert add(conn, article("https://b.test/2", "حمله موشکی به تاسیسات"))[2] is True

    @pytest.mark.parametrize("left,right", [
        # Real pairs from the production corpus, all scoring >= 0.70.
        ("گسترده ترین حمله موشکی ایران در هفته های اخیر | انفجار در شهرک",
         "گسترده ترین حمله موشکی ایران در هفته های اخیر؛ انفجار در شهرک"),
        ("رادارهای اولیه این شکلی بودند", "رادارهای اولیه این شکلی بودند/ عکس"),
        ("چگونه در بحران، آرامش خود را حفظ کنیم؟", "چگونه در بحران آرامش خود را حفظ کنیم؟"),
        ("نفتکش حامل نفت خام ایران از تنگه هرمز عبور کرد/ عکس",
         "نفتکش حامل نفت خام ایران از تنگه هرمز عبور کرد"),
    ])
    def test_reworded_titles_are_merged(self, conn, left, right):
        register_sources(conn)
        add(conn, article("https://a.test/1", left, content="متن یک"))
        assert add(conn, article("https://b.test/2", right, content="متن دو"))[2] is True
        assert canonical_count(conn) == 1

    @pytest.mark.parametrize("left,right", [
        # Also real pairs, scoring 0.5-0.62: different articles sharing a date template.
        ("فال قهوه سه شنبه یک اردیبهشت ۱۴۰۵", "فال روزانه سه شنبه یک اردیبهشت 1405"),
        ("فال انبیا چهارشنبه ۲ اردیبهشت ۱۴۰۵", "فال روزانه چهارشنبه 2 اردیبهشت ۱۴۰۵"),
        ("صفحه نخست روزنامه ها چهارشنبه ۲ اردیبهشت ۱۴۰۵", "فال انبیا چهارشنبه ۲ اردیبهشت ۱۴۰۵"),
        # The one false positive at 0.704: two different cities sharing the بندر stem.
        # This pair is why the threshold sits at 0.75 rather than 0.70.
        ("احتمال شنیده شدن صدای انفجارهای کنترل شده در شرق شهر بندرعباس",
         "احتمال شنیده شدن صدای انفجارهای کنترل شده در بندرلنگه"),
        # Genuinely unrelated stories.
        ("افزایش قیمت طلا در بازار تهران", "برگزاری مسابقات فوتبال جوانان"),
    ])
    def test_similar_but_distinct_titles_are_not_merged(self, conn, left, right):
        register_sources(conn)
        add(conn, article("https://a.test/1", left, content="متن یک"))
        assert add(conn, article("https://b.test/2", right, content="متن دو"))[2] is False
        assert canonical_count(conn) == 2


class TestTimeWindowBlocking:
    def test_the_same_headline_weeks_apart_is_a_new_story(self, conn):
        register_sources(conn)
        add(conn, article("https://a.test/1", "وضعیت بازار ارز امروز", content="a",
                          published="2026-08-01T10:00:00+03:30"))
        assert add(conn, article("https://b.test/2", "وضعیت بازار ارز امروز", content="b",
                                 published="2026-08-20T10:00:00+03:30"))[2] is False

    def test_undated_near_duplicates_are_still_caught(self, conn):
        """Undated articles used to return zero candidates and skip near-duplicate
        detection entirely; shahrekhabar produces them routinely."""
        register_sources(conn)
        add(conn, article("https://a/1", "رادارهای اولیه این شکلی بودند",
                          source="shahrekhabar", published=None))
        assert add(conn, article("https://a/2", "رادارهای اولیه این شکلی بودند/ عکس",
                                 source="shahrekhabar", content="متن دیگر", published=None))[2] is True

    def test_undated_candidates_are_bounded(self, conn):
        assert len(dedupe.candidates(conn, article_id=-1, published_at=None)) <= \
            dedupe.UNDATED_CANDIDATE_LIMIT

    def test_candidates_are_limited_to_the_window(self, conn):
        register_sources(conn)
        # Titles must be genuinely unrelated, or they dedupe into each other and the
        # candidate pool is 1. Numbered variants of one sentence are near-identical.
        topics = ["طلا", "نفت", "گندم", "مسکن", "خودرو", "بورس", "ارز", "بیمه", "مالیات", "بودجه"]
        for offset, when in ((0, "2026-08-16T10:00:00+03:30"), (100, "2026-01-01T10:00:00+03:30")):
            for i, topic in enumerate(topics, start=offset):
                add(conn, article(f"https://a.test/{i}",
                                  f"گزارش کامل درباره وضعیت {topic} در کشور",
                                  content=f"متن {i}", published=when))
        near = dedupe.candidates(conn, article_id=-1, published_at="2026-08-16T10:00:00+03:30")
        assert len(near) == len(topics), "articles months away must not be compared"


class TestCanonicalSelection:
    def test_higher_priority_source_becomes_canonical(self, conn):
        """Mehr arrives first, Khabarfoori second with a fuller body: Khabarfoori wins."""
        register_sources(conn)
        first = add(conn, article("https://mehr.test/1", source="mehr", content="خلاصه کوتاه"))[0]
        second = add(conn, article("https://kf.test/2", content="متن کامل خبر"))[0]
        rows = duplicate_map(conn)
        assert rows[second] is None and rows[first] == second

    def test_lower_priority_arrival_does_not_displace(self, conn):
        register_sources(conn)
        first = add(conn, article("https://kf.test/1", content="متن کامل"))[0]
        second = add(conn, article("https://sk.test/2", source="shahrekhabar", content=""))[0]
        rows = duplicate_map(conn)
        assert rows[first] is None and rows[second] == first

    def test_longer_content_wins_within_the_same_source(self, conn):
        register_sources(conn)
        first = add(conn, article("https://a.test/1", content="کوتاه"))[0]
        second = add(conn, article("https://a.test/2", content="متن بسیار طولانی تر از نسخه اول"))[0]
        rows = duplicate_map(conn)
        assert rows[second] is None and rows[first] == second

    def test_an_already_classified_canonical_is_not_displaced(self, conn):
        """A duplicate discovered later (e.g. by backfill, which always runs the free rule
        provider) must not demote an article carrying a real classification - that would
        orphan the inference behind a `duplicate_of` row no view looks at again."""
        register_sources(conn)
        first = add(conn, article("https://sk.test/1", source="shahrekhabar", content="کوتاه"))[0]
        db.insert(conn, "classifications", {
            "article_id": first, "category": "security/economics", "confidence": "زیاد",
            "method": "llm", "prompt_version": "v1", "provider": "gapgpt", "model": "m1",
            "run_id": "r1", "created_at": "2026-08-16T10:00:00+00:00",
        })
        second = add(conn, article("https://kf.test/2", content="متن کامل خبر بسیار طولانی تر"))[0]
        rows = duplicate_map(conn)
        assert rows[first] is None, "already-classified article must stay canonical"
        assert rows[second] == first

    def test_a_duplicate_chain_never_exceeds_one_level(self, conn):
        """Demoting a canonical must repoint its followers, or `duplicate_of IS NULL`
        stops being a reliable "this is the story" filter."""
        register_sources(conn)
        a = add(conn, article("https://sk.test/1", source="shahrekhabar", content=""))[0]
        b = add(conn, article("https://mehr.test/2", source="mehr", content="خلاصه"))[0]
        c = add(conn, article("https://kf.test/3", content="متن کامل خبر"))[0]
        rows = duplicate_map(conn)
        assert rows[c] is None
        assert rows[a] == c and rows[b] == c, "no duplicate may point at another duplicate"
        assert canonical_count(conn) == 1


class TestDedupeBackfill:
    @pytest.fixture
    def unlinked_pair(self, conn):
        register_sources(conn)
        add(conn, article("https://a.test/1", "رادارهای اولیه این شکلی بودند", content="a"))
        add(conn, article("https://b.test/2", "رادارهای اولیه این شکلی بودند/ عکس", content="b"))
        conn.execute("UPDATE articles SET duplicate_of = NULL")  # undo ingest-time linking
        return conn

    def test_dry_run_reports_each_pair_once_and_changes_nothing(self, unlinked_pair):
        """Without linking there is nothing to suppress the reverse match, so the naive
        version reports A->B and B->A and doubles the count."""
        assert len(dedupe.backfill(unlinked_pair, dry_run=True)) == 1
        assert canonical_count(unlinked_pair) == 2

    def test_apply_links_the_matches_and_is_idempotent(self, unlinked_pair):
        assert len(dedupe.backfill(unlinked_pair)) == 1
        assert canonical_count(unlinked_pair) == 1
        assert dedupe.backfill(unlinked_pair) == [], "a second pass must find nothing"

    def test_find_duplicate_returns_none_when_unique(self, conn):
        register_sources(conn)
        article_id = add(conn, article("https://a.test/1", "یک خبر کاملا منحصر به فرد"))[0]
        assert dedupe.find_duplicate(conn, article_id=article_id, title="عنوان بی ربط دیگر",
                                     content_hash="nothing",
                                     published_at="2026-08-16T10:00:00+03:30") is None

    def test_an_empty_title_is_not_matched(self, conn):
        assert dedupe.find_duplicate(conn, article_id=1, title="", content_hash="x",
                                     published_at="2026-08-16T10:00:00+03:30") is None


# ---------------------------------------------------------------- window backfill


def jalali_today(offset=0):
    from datetime import timedelta as delta
    day = jdatetime.date.today() - delta(days=offset)
    return f"{day.year:04d}-{day.month:02d}-{day.day:02d}"


def dated_article(conn, *, source, persian_date, uncertain=0):
    url = f"https://test/{source}/{persian_date}/{uncertain}"
    db.insert(conn, "articles", {
        "url": url, "identity_key": f"id:{url}", "source": source, "original_title": "t",
        "lead": "l", "content": "c", "content_hash": url, "published_at_persian": persian_date,
        "date_uncertain": uncertain, "fetched_at": "2026-01-01T00:00:00+03:30",
    })


class TestWindowBackfill:
    def test_missing_days_reports_every_gap_in_the_window(self, conn):
        dated_article(conn, source="khabarfoori", persian_date=jalali_today())
        gaps = pipeline.missing_days(conn, "khabarfoori", days=3)
        assert len(gaps) == 2 and jalali_today() not in gaps

    def test_uncertain_dates_do_not_count_as_coverage(self, conn):
        dated_article(conn, source="khabarfoori", persian_date=jalali_today(), uncertain=1)
        assert jalali_today() in pipeline.missing_days(conn, "khabarfoori", days=1)

    def test_coverage_is_full_once_every_window_day_has_an_article(self, conn):
        for offset in range(3):
            dated_article(conn, source="mehr", persian_date=jalali_today(offset))
        assert pipeline.missing_days(conn, "mehr", days=3) == set()

    @pytest.mark.parametrize("spec,days", [
        # No gap -> never attempted, so no network or session was ever needed.
        (SourceSpec("khabarfoori", 2, "listing_detail", "https://kf.test"), 2),
        # No history mechanism, and disabled sources, are both skipped outright.
        (SourceSpec("shahrekhabar", 2, "listing_relay", "https://shahr.test"), 14),
        (SourceSpec("khabarfoori", 2, "listing_detail", "https://kf.test", enabled=False), 14),
    ])
    def test_ensure_window_skips_what_it_cannot_or_need_not_fetch(self, conn, spec, days):
        if days == 2:
            for offset in range(2):
                dated_article(conn, source="khabarfoori", persian_date=jalali_today(offset))
        assert pipeline.ensure_window(conn, {spec.name: spec}, {}, days=days) == {}

    def test_ensure_window_respects_the_retry_cooldown(self, conn, monkeypatch):
        """A gap that did not close on the last attempt is not retried immediately."""
        calls = []
        monkeypatch.setattr(pipeline.sources, "backfill_fetch",
                            lambda spec, session=None, **kw: iter(calls.append(1) or ()))
        specs = {"khabarfoori": SourceSpec("khabarfoori", 2, "listing_detail", "https://kf.test")}
        assert pipeline.ensure_window(conn, specs, {}, days=14) == {"khabarfoori": 0}
        assert pipeline.ensure_window(conn, specs, {}, days=14) == {}, "cooldown active"
        assert len(calls) == 1


# --------------------------------------------------------------------------- export


def test_export_preserves_the_workbook_template(conn, tmp_path):
    pipeline.process(conn, [article()], RuleProvider(), run_id="run")
    result = exports.export_all(conn, tmp_path / "out")
    assert all(path.exists() for path in result.values())

    from openpyxl import load_workbook
    workbook_path = next(path for name, path in result.items() if name.startswith("excel:"))
    template = load_workbook(config.WORKBOOK_TEMPLATE_PATH)
    workbook = load_workbook(workbook_path, data_only=False)
    sheet = workbook["بررسی خبر"]

    # The daily file is the single operational sheet, matching every workbook the team
    # produced. The template's reference sheets stay in the template.
    assert workbook.sheetnames == ["بررسی خبر"]
    assert sheet["A2"].style_id == template["بررسی خبر"]["A2"].style_id
    assert str(sheet["J3"].value).startswith("=IF(")
    # The template's own header row leaves L2 blank; the export has to supply it or the
    # link column ships unlabelled, as it did in every legacy workbook.
    assert [sheet.cell(2, column).value for column in range(1, 13)] == exports.HEADERS

    validations = sheet.data_validations.dataValidation
    trend = next(v for v in validations if v.formula1.startswith('"↑'))
    levels = next(v for v in validations if v is not trend)
    values = lambda v: v.formula1.strip('"').split(",")
    assert values(trend) == list(prompts.GOLD_TRENDS)
    assert values(levels) == list(LEVELS), "score columns are ordinal, not yes/no"

    with zipfile.ZipFile(config.WORKBOOK_TEMPLATE_PATH) as source, \
         zipfile.ZipFile(workbook_path) as output:
        extension = lambda xml: xml[xml.find(b"<extLst"):xml.find(b"</extLst>") + len(b"</extLst>")]
        assert extension(output.read("xl/worksheets/sheet1.xml")) == \
            extension(source.read("xl/worksheets/sheet1.xml"))


# ------------------------------------------------------------------------------ cli


def test_backfilled_articles_always_classify_on_rule_never_the_runs_real_provider(
    tmp_path, monkeypatch
):
    """A coverage gap can mean hundreds of articles - paying to label all of them as a
    silent side effect of a routine `run --provider gapgpt` would be a real-money surprise,
    not something the $1 budget ceiling catching it after the fact makes acceptable."""
    monkeypatch.setattr(config, "DB_PATH", tmp_path / "news.db")
    monkeypatch.setattr(config, "ensure_dirs", lambda: None)
    monkeypatch.setattr(providers, "resolve", lambda choice: {
        node: RuleProvider() if choice == "rule" else RuleProvider(name="real-provider", model="m1")
        for node in pipeline.NODES
    })

    body = "متن کامل خبر درباره حمله و اثر آن بر قیمت طلا و دلار در بازار داخلی."
    monkeypatch.setattr(sources, "fetch", lambda spec, session=None, *, limit=25: [RawArticle(
        source=spec.name, url=f"https://test/{spec.name}/main",
        title="حمله موشکی به تاسیسات نفتی کشور", lead="جزئیات حادثه", content=body,
        published_at=datetime.now(timezone.utc).isoformat())])

    def fake_backfill(spec, session=None, *, since_date, known_urls):
        if spec.name == "khabarfoori":
            yield RawArticle(
                source="khabarfoori", url="https://test/khabarfoori/backfilled",
                title="افزایش قیمت طلا و نگرانی امنیتی در بازار", lead="جزئیات", content=body,
                published_at=(datetime.now(timezone.utc) - timedelta(days=5)).isoformat())

    monkeypatch.setattr(sources, "backfill_fetch", fake_backfill)
    cli.run_once(argparse.Namespace(sources=None, limit=5, provider="gapgpt", export=False))

    with db.connect(tmp_path / "news.db", readonly=True) as conn:
        by_url = {row["url"]: row["provider"] for row in conn.execute(
            "SELECT a.url AS url, c.provider AS provider"
            " FROM articles a JOIN classifications c ON c.article_id = a.id")}
    assert by_url["https://test/khabarfoori/main"] == "real-provider"
    assert by_url["https://test/khabarfoori/backfilled"] == "rule"


def test_the_cli_exposes_every_documented_command():
    parser = cli.build_parser(["khabarfoori"])
    commands = next(a for a in parser._actions if a.dest == "command").choices
    assert set(commands) == {
        "init", "run", "run-loop", "replay", "export", "routes", "dedupe", "review-queue",
        "review-import", "canary", "evaluate", "golden", "compare", "serve",
    }


def test_run_loop_survives_a_failing_cycle_but_stops_on_fatal(monkeypatch):
    attempts = []

    def failing(args):
        attempts.append(1)
        raise RuntimeError("one bad cycle")

    monkeypatch.setattr(cli, "run_once", failing)
    monkeypatch.setattr(cli.time, "sleep", lambda _: None)
    args = argparse.Namespace(interval_minutes=1, limit=1, provider="rule", export=False)
    assert cli.run_loop(args, cycles=2) == 0
    assert len(attempts) == 2, "a bad cycle must not end the daemon"

    monkeypatch.setattr(cli, "run_once", lambda args: (_ for _ in ()).throw(dag.Fatal("bad key")))
    assert cli.run_loop(args, cycles=2) == 1, "Fatal stops the loop"
